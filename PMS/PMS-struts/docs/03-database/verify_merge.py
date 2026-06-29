#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证合并结果：检查pm_project表的字段顺序和新列数据"""
import openpyxl

wb = openpyxl.load_workbook(r'd:\EclipseWorkspace\Parctice\PMS\PMS-struts\docs\03-database\database_dict_flat_final_merged_v2.xlsx')
ws = wb['项目管理']

# 找到pm_project表的数据
in_table = False
count = 0
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    c = ws.cell(r, 3).value

    if a and str(a).strip() == 'pm_project':
        in_table = True
    elif a and str(a).strip() != 'pm_project' and in_table:
        break

    if in_table and c:
        fname = str(c).strip()
        comment = str(ws.cell(r, 8).value or '').strip()
        biz = str(ws.cell(r, 9).value or '').strip()
        source = str(ws.cell(r, 10).value or '').strip()  # 来源
        biz_match = str(ws.cell(r, 11).value or '').strip()  # 对应业务
        data_asset = str(ws.cell(r, 14).value or '').strip()  # 数据资产
        app_theme = str(ws.cell(r, 16).value or '').strip()  # 应用主题

        new_cols = []
        if source: new_cols.append(f"来源={source}")
        if biz_match: new_cols.append(f"对应业务={biz_match}")
        if data_asset: new_cols.append(f"数据资产={data_asset}")
        if app_theme: new_cols.append(f"应用主题={app_theme}")

        new_info = ' | '.join(new_cols) if new_cols else ''
        print(f"  {fname:30s} | {comment[:20]:20s} | {new_info}")
        count += 1

print(f"\npm_project 总字段数: {count}")
