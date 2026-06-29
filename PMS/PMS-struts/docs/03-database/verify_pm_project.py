#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""对比pm_project在匹配表和基表中的字段"""
import openpyxl, re

# 匹配表
wb2 = openpyxl.load_workbook(r'd:\EclipseWorkspace\Parctice\PMS\PMS-struts\docs\03-database\项目管理表数据元.xlsx')
ws2 = wb2[wb2.sheetnames[0]]

# 找到pm_project的匹配数据
in_table = False
match_fields = []
for r in range(1, ws2.max_row + 1):
    c1 = ws2.cell(r, 1).value
    c2 = ws2.cell(r, 2).value
    c6 = ws2.cell(r, 6).value

    if c1 and str(c1).strip() == 'pm_project' and c2 is None:
        in_table = True
        continue
    if in_table and c1 and str(c1).strip() not in ('字段名', '对象类型', '业务含义', '数据量', '数据大小') and '——' not in str(c1):
        if re.match(r'^[a-zA-Z_]', str(c1).strip()):
            in_table = False
            continue
    if in_table and c1 and str(c1).strip() == '字段名':
        continue
    if in_table and c1 and str(c1).strip() in ('对象类型', '业务含义', '数据量', '数据大小', '索引列表'):
        continue

    if in_table:
        fname = str(c1).strip() if c1 else ''
        comment = str(c6).strip() if c6 else ''
        if fname or comment:
            match_fields.append({'fname': fname, 'comment': comment})

# 基表
wb1 = openpyxl.load_workbook(r'd:\EclipseWorkspace\Parctice\PMS\PMS-struts\docs\03-database\database_dict_flat_final_match.xlsx')
ws1 = wb1['项目管理']
base_fields = []
for r in range(1, ws1.max_row + 1):
    a = ws1.cell(r, 1).value
    c = ws1.cell(r, 3).value
    if a and str(a).strip() == 'pm_project' and c:
        base_fields.append(str(c).strip())

base_set = set(base_fields)

print("匹配表 pm_project 字段列表:")
for i, mf in enumerate(match_fields):
    fname = mf['fname']
    comment = mf['comment']
    if fname and fname in base_set:
        status = "匹配"
    elif fname:
        status = "新增(有字段名)"
    else:
        status = "新增(无字段名)"
    print(f"  {i+1:2d}. fname='{fname:25s}' comment='{comment:20s}' -> {status}")

print(f"\n匹配表字段数: {len(match_fields)}")
print(f"基表字段数: {len(base_fields)}")
