#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""详细检查匹配表中表名行和字段行的结构"""
import openpyxl
import re

wb2 = openpyxl.load_workbook(r'd:\EclipseWorkspace\Parctice\PMS\PMS-struts\docs\03-database\项目管理表数据元.xlsx')
ws = wb2[wb2.sheetnames[0]]

# 找出所有表名行和字段数据行
table_count = 0
field_count = 0
table_names = []

for r in range(1, ws.max_row+1):
    c1 = ws.cell(r, 1).value
    c6 = ws.cell(r, 6).value
    c2 = ws.cell(r, 2).value

    # 表名行：C1有值且包含中文或——，C2为None
    if c1 and c2 is None and '——' in str(c1):
        table_count += 1
        # 解析表名
        parts = str(c1).split('——')
        tname = parts[0].strip()
        tdesc = parts[1].strip() if len(parts) > 1 else ''
        table_names.append(tname)
        if table_count <= 5:
            print(f"  表名行 Row {r}: tname='{tname}' tdesc='{tdesc}'")

    # 字段数据行：C6有值且C2为None（字段描述列有值，但字段名列为空）
    elif c6 and c2 is None and str(c1) != '字段名':
        field_count += 1
        if field_count <= 10:
            c8 = ws.cell(r, 8).value
            c9 = ws.cell(r, 9).value
            c10 = ws.cell(r, 10).value
            c11 = ws.cell(r, 11).value
            c12 = ws.cell(r, 12).value
            c13 = ws.cell(r, 13).value
            c14 = ws.cell(r, 14).value
            c15 = ws.cell(r, 15).value
            c16 = ws.cell(r, 16).value
            print(f"  字段行 Row {r}: desc='{c6}' | 来源={c8} 对应业务={c9} 数据元类型={c10} 数据现状={c11} 处理方案={c12} 数据资产={c13} 关联资产={c14} 应用主题={c15} 自动回传={c16}")

print(f"\n总表数: {table_count}")
print(f"总字段行数: {field_count}")
print(f"\n所有表名: {table_names[:20]}")
