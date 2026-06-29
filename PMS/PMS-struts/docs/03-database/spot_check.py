#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""对比 MD 文件和 Excel 中特定表的字段数"""
import re, os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))

# 检查几个关键表
CHECK_TABLES = ['pm_project', 'pm_project_header_view_cache', 'act_ru_execution', 'fb_contract', 'rma_applicant']

def read_md(filepath):
    with open(filepath, 'rb') as f:
        data = f.read()
    return data.decode('utf-8', errors='replace').replace('\r\n','\n').replace('\r','\n')

# 从 MD 统计
text = read_md(os.path.join(BASE, 'database_dict final.md'))
lines = text.split('\n')

for tname in CHECK_TABLES:
    # 找到表定义
    for i, line in enumerate(lines):
        m = re.match(r'^#{1,5}\s+[\d.]*\s*' + re.escape(tname) + r'\b', line.strip())
        if m:
            # 数字段行
            field_count = 0
            j = i + 1
            while j < len(lines):
                l = lines[j].strip()
                if l.startswith('| 字段名'):
                    j += 1; j += 1
                    while j < len(lines) and lines[j].strip().startswith('|') and '---' not in lines[j]:
                        field_count += 1
                        j += 1
                    break
                j += 1
            print(f"MD: {tname} = {field_count} 字段")
            break

# 从 Excel 统计
XLSX_FILE = os.path.join(BASE, 'database_dict_final_v2.xlsx')
wb = load_workbook(XLSX_FILE, read_only=True)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    in_fields = False
    current_table = ''
    field_count = 0

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        val0 = str(row[0]).strip() if row[0] else ''

        if '——' in val0:
            # 保存上一个表的字段数
            if current_table in CHECK_TABLES:
                print(f"Excel: {current_table} = {field_count} 字段")
            current_table = val0.split('——')[0].strip()
            field_count = 0
            in_fields = False
            continue

        if val0 == '字段名':
            in_fields = True
            continue

        if val0 in ('索引名', '外键名', '索引列表', '外键列表'):
            in_fields = False
            continue

        if in_fields and row[1]:
            dtype = str(row[1]).strip().lower()
            if any(t in dtype for t in ['varchar', 'int', 'datetime', 'text', 'decimal', 'bigint', 'tinyint', 'date', 'timestamp', 'char', 'double', 'float', 'longtext', 'mediumtext', 'smallint', 'blob', 'enum', 'bit', 'time']):
                field_count += 1

    if current_table in CHECK_TABLES:
        print(f"Excel: {current_table} = {field_count} 字段")

wb.close()
