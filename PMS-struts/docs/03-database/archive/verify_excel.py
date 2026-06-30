#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证 Excel 导出的字段描述和业务含义完整性"""
from openpyxl import load_workbook
import os

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE = os.path.join(BASE, 'database_dict_final_v2.xlsx')

wb = load_workbook(XLSX_FILE, read_only=True)

total_tables = 0
total_fields = 0
empty_comment = 0
empty_biz = 0
empty_both = 0

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    table_count = 0
    for row in ws.iter_rows(min_row=1, values_only=False):
        vals = [c.value for c in row]
        # 检测表名标题行（深蓝背景）
        if vals[0] and isinstance(vals[0], str) and '——' in str(vals[0]):
            table_count += 1
        # 检测字段数据行（第1列是字段名，不是表头）
        if vals[0] and isinstance(vals[0], str) and vals[0] not in ('字段名', '索引名', '外键名', '索引列表', '外键列表', '对象类型', '业务含义', '数据量', '数据大小', '属性', None):
            # 排除标题行和属性行
            first = str(vals[0]).strip()
            if first and not first.startswith('——') and not first.startswith('|') and len(first) < 50:
                # 检查是否是字段数据行（有数据类型列）
                if vals[1] and isinstance(vals[1], str) and ('varchar' in str(vals[1]).lower() or 'int' in str(vals[1]).lower() or 'datetime' in str(vals[1]).lower() or 'text' in str(vals[1]).lower() or 'decimal' in str(vals[1]).lower() or 'bigint' in str(vals[1]).lower() or 'tinyint' in str(vals[1]).lower() or 'date' in str(vals[1]).lower() or 'timestamp' in str(vals[1]).lower() or 'char' in str(vals[1]).lower() or 'double' in str(vals[1]).lower() or 'float' in str(vals[1]).lower() or 'longtext' in str(vals[1]).lower() or 'mediumtext' in str(vals[1]).lower() or 'smallint' in str(vals[1]).lower() or 'blob' in str(vals[1]).lower() or 'enum' in str(vals[1]).lower() or 'bit' in str(vals[1]).lower() or 'time' in str(vals[1]).lower()):
                    total_fields += 1
                    comment = str(vals[5]).strip() if vals[5] else ''
                    biz = str(vals[6]).strip() if vals[6] else ''
                    if not comment or comment == '-' or comment == 'None':
                        empty_comment += 1
                    if not biz or biz == '-' or biz == 'None' or biz == '业务含义待确认':
                        empty_biz += 1
                    if (not comment or comment == '-' or comment == 'None') and (not biz or biz == '-' or biz == 'None' or biz == '业务含义待确认'):
                        empty_both += 1
    total_tables += table_count
    print(f"  {ws_name}: {table_count} 张表")

wb.close()

print(f"\n总表数: {total_tables}")
print(f"总字段数: {total_fields}")
print(f"字段描述为空: {empty_comment} ({empty_comment*100//total_fields if total_fields else 0}%)")
print(f"业务含义为空/待确认: {empty_biz} ({empty_biz*100//total_fields if total_fields else 0}%)")
print(f"两者都为空: {empty_both}")
