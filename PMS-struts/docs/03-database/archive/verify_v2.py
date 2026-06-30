#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""直接对比 MD 解析结果与 Excel 导出结果"""
import re, os
from openpyxl import load_workbook

BASE = os.path.dirname(os.path.abspath(__file__))

def read_md(filepath):
    with open(filepath, 'rb') as f:
        data = f.read()
    return data.decode('utf-8', errors='replace').replace('\r\n','\n').replace('\r','\n')

def parse_all_tables(text):
    """完整解析MD中所有表的字段"""
    tables = {}
    lines = text.split('\n')
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        m = re.match(r'^#{1,5}\s+[\d.]*\s*(\S+)\s*(?:--\s*(.+))?', line)
        if not m or not m.group(1) or len(m.group(1))<2 or re.search(r'[\u4e00-\u9fff]',m.group(1)) or m.group(1).startswith('**') or m.group(1).startswith('第') or m.group(1).startswith('附录') or m.group(1)=='DPPMS':
            i += 1
            continue
        tname = m.group(1)
        obj_type = ''
        fields = []
        i += 1
        while i < len(lines):
            l = lines[i].strip()
            if l.startswith('| 对象类型'):
                parts = [p.strip() for p in l.split('|')]
                parts = [p for p in parts if p]
                if len(parts) >= 2:
                    obj_type = parts[1]
            if l.startswith('| 字段名'):
                i += 1; i += 1
                while i < len(lines) and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    raw = lines[i].strip()
                    parts = [p.strip() for p in raw.split('|')]
                    parts = [p for p in parts if p]
                    if len(parts) >= 1:
                        fname = parts[0]
                        comment = parts[5] if len(parts) >= 6 else ''
                        biz = parts[6] if len(parts) >= 7 else ''
                        fields.append({'name': fname, 'comment': comment, 'biz': biz, 'col_count': len(parts)})
                    i += 1
                break
            elif l.startswith('**约束') or l.startswith('**索引') or l.startswith('**外键') or l=='---' or re.match(r'^#{1,5}',l):
                break
            else:
                i += 1
        tables[tname] = {'obj_type': obj_type, 'fields': fields}
        while i < len(lines):
            l = lines[i].strip()
            if l=='---' or re.match(r'^#{1,5}\s+',l):
                break
            i += 1
    return tables

# 解析 MD
text = read_md(os.path.join(BASE, 'database_dict final.md'))
md_tables = parse_all_tables(text)

# 过滤 temp/tmp/视图
filtered = {k: v for k, v in md_tables.items() if not k.startswith(('temp_','tmp_')) and v['obj_type'] != 'VIEW'}

# 统计
total_fields = 0
empty_comment = 0
empty_biz = 0
col_dist = {}
for tname, data in sorted(filtered.items()):
    for f in data['fields']:
        total_fields += 1
        col_dist[f['col_count']] = col_dist.get(f['col_count'], 0) + 1
        if not f['comment'] or f['comment'] == '-':
            empty_comment += 1
        if not f['biz'] or f['biz'] == '-' or f['biz'] == '业务含义待确认':
            empty_biz += 1

print(f"MD文件统计（过滤后）:")
print(f"  表数: {len(filtered)}")
print(f"  字段数: {total_fields}")
print(f"  列数分布: {col_dist}")
print(f"  字段描述为空: {empty_comment} ({empty_comment*100//total_fields}%)")
print(f"  业务含义为空/待确认: {empty_biz} ({empty_biz*100//total_fields}%)")

# 检查非7列表
non7 = [(tname, f) for tname, data in filtered.items() for f in data['fields'] if f['col_count'] != 7]
if non7:
    print(f"\n非7列字段 ({len(non7)} 个):")
    for tname, f in non7[:10]:
        print(f"  {tname}.{f['name']}: {f['col_count']}列, comment='{f['comment'][:30]}', biz='{f['biz'][:30]}'")

# 检查 Excel
print("\n\nExcel验证:")
XLSX_FILE = os.path.join(BASE, 'database_dict_final_v2.xlsx')
wb = load_workbook(XLSX_FILE, read_only=True)
for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    # 统计字段表头行
    field_count = 0
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row[0] == '字段名':
            field_count += 1
    print(f"  {ws_name}: 字段表头 {field_count} 个")
wb.close()
