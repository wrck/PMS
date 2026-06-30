#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""检查两个Excel文件的结构"""
import openpyxl

# 基表
wb1 = openpyxl.load_workbook(r'd:\EclipseWorkspace\Parctice\PMS\PMS-struts\docs\03-database\database_dict_flat_final_match.xlsx')
print("=== 基表: database_dict_flat_final_match.xlsx ===")
for sn in wb1.sheetnames:
    ws = wb1[sn]
    print(f"\nSheet: {sn}, rows={ws.max_row}, cols={ws.max_column}")
    # 打印前3行
    for r in range(1, min(4, ws.max_row+1)):
        row_data = []
        for c in range(1, ws.max_column+1):
            row_data.append(str(ws.cell(r, c).value)[:30])
        print(f"  Row {r}: {row_data}")

# 匹配表
wb2 = openpyxl.load_workbook(r'd:\EclipseWorkspace\Parctice\PMS\PMS-struts\docs\03-database\项目管理表数据元.xlsx')
print("\n\n=== 匹配表: 项目管理表数据元.xlsx ===")
for sn in wb2.sheetnames:
    ws = wb2[sn]
    print(f"\nSheet: {sn}, rows={ws.max_row}, cols={ws.max_column}")
    # 打印前5行
    for r in range(1, min(6, ws.max_row+1)):
        row_data = []
        for c in range(1, ws.max_column+1):
            row_data.append(str(ws.cell(r, c).value)[:30])
        print(f"  Row {r}: {row_data}")
