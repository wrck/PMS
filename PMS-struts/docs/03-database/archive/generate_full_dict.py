#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成完整数据字典（MD + Excel）
- 数据源：数据库schema JSON + 现有MD业务含义 + 代码业务含义
- 包含：所有表/视图、字段、约束（PK/FK/UNIQUE/其他）、索引
- 保留已有业务含义，用代码含义补全缺失部分
"""
import json
import os
import re
import copy
import pymysql
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SCHEMA_DIR = os.path.join(BASE_DIR, 'schema_data')
MD_OUT = os.path.join(BASE_DIR, 'database_dict_full.md')
XLSX_OUT = os.path.join(BASE_DIR, 'database_dict_full.xlsx')

# ============================================================
# 业务域分类规则
# ============================================================
DOMAIN_RULES = [
    ("项目管理", [
        # 一、项目管理域 - pm_project* 系列（前缀匹配）
        r'^pm_project_',
        r'^pm_project$',
        r'^pm_basic',
        r'^pm_column',
        r'^pm_common',
        r'^pm_daily',
        r'^pm_dispatch',
        r'^pm_notification',
        r'^pm_product_info',
        r'^pm_facilitator',
        # 二、回访管理域
        r'^pm_cl_',
        # 三、售前管理域
        r'^pm_presales',
        # 四、转包管理域
        r'^pm_subcontract',
        r'^pm_sub_',
        # 五、问题管理域
        r'^prob_',
        # 六、基础平台域
        r'^fnd_',
        # 无前缀的项目管理相关表
        r'^(addressee_info|agent_info|back_type|serve_type|tain_type)$',
    ]),
    ("系统支撑", [
        # 一、EHR组织架构域
        r'^ehr_',
        # 二、系统权限域
        r'^t_',
        r'^dp_',
        r'^(role|user|user_info|user_modules|user_permissions|user_team|tb_sys_log)$',
        # 三、数据同步中间表域
        r'^pm_order_data',
        r'^pm_order_line',
        r'^pm_pb_plan',
        r'^pm_person_from',
        r'^pm_presales_lend',
        r'^pm_project_property_from',
        r'^pm_project_property_af',
        r'^pm_project_real_product',
        r'^pm_project_product_af',
        r'^pm_project_soleagent_lend',
        r'^pm_project_market_relations',
        r'^project_info_from',
        r'^pm_project_product_config',
        r'^pm_project_product_lease',
        r'^pm_project_incident_table',
        r'^sms_ofst_contract',
        # 四、其他辅助表
        r'^pm_report',
        r'^pm_workflow',
        r'^pm_data_refresh',
        r'^(data_field_relation|hexiao|transnum|sys_state_or_type|firebird_operation_log)$',
    ]),
    ("历史迁移与引擎", [
        # 一、Activiti工作流引擎表
        r'^act_',
        r'^fnd_act_',
        # 二、Firebird迁移表
        r'^fb_',
        # 三、RMA/备件/仓库等业务表
        r'^rma_',
        r'^warehouse',
        r'^spare_',
        r'^department$',
        # 无前缀的迁移/业务表
        r'^af_industry_',
        r'^brw_',
        r'^app_',
        r'^mes_',
        r'^warranty_',
        r'^(tx_info|bar|workflow_info|find_in_set_help)$',
        r'^shipment_barcode',
        r'^dptech_v_',
        r'^view_warranty',
    ]),
]

# 视图单独一个sheet
VIEW_DOMAIN = "视图"


def classify_table(table_name, table_type):
    """根据表名分类到业务域"""
    if table_type == 'VIEW':
        return VIEW_DOMAIN
    if table_name.startswith(('temp_', 'tmp_')):
        return None  # 过滤临时表
    for domain, patterns in DOMAIN_RULES:
        for pattern in patterns:
            if re.match(pattern, table_name):
                return domain
    return "其他"


def load_json(filename):
    filepath = os.path.join(SCHEMA_DIR, filename)
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def build_data():
    """构建完整数据结构"""
    # 1. 加载数据
    objects = load_json('objects.json')
    columns = load_json('columns.json')
    constraints = load_json('constraints.json')
    indexes = load_json('indexes.json')
    foreign_keys = load_json('foreign_keys.json')
    existing_biz = load_json('existing_biz_meanings.json')
    code_biz = load_json('code_biz_meanings.json')

    # 2. 构建索引
    # 按表名组织字段
    cols_by_table = defaultdict(list)
    for col in columns:
        cols_by_table[col['TABLE_NAME']].append(col)

    # 按表名组织约束
    constraints_by_table = defaultdict(lambda: defaultdict(list))
    for c in constraints:
        tname = c['TABLE_NAME']
        ctype = c['CONSTRAINT_TYPE']
        constraints_by_table[tname][ctype].append(c)

    # 按表名组织索引
    idx_by_table = defaultdict(list)
    idx_grouped = defaultdict(lambda: defaultdict(list))
    for idx in indexes:
        tname = idx['TABLE_NAME']
        idx_name = idx['INDEX_NAME']
        idx_grouped[tname][idx_name].append(idx)
    for tname, idx_groups in idx_grouped.items():
        for idx_name, idx_cols in idx_groups.items():
            first = idx_cols[0]
            col_list = ', '.join(c['COLUMN_NAME'] for c in sorted(idx_cols, key=lambda x: x['SEQ_IN_INDEX']))
            idx_by_table[tname].append({
                'INDEX_NAME': idx_name,
                'INDEX_TYPE': first['INDEX_TYPE'],
                'NON_UNIQUE': first['NON_UNIQUE'],
                'COLUMNS': col_list,
            })

    # 按表名组织外键
    fk_by_table = defaultdict(list)
    for fk in foreign_keys:
        fk_by_table[fk['TABLE_NAME']].append(fk)

    # 3. 合并业务含义
    def get_field_biz_meaning(table_name, field_name, db_comment=''):
        """获取字段业务含义，优先级：现有MD > 代码 > 数据库注释"""
        # 1. 现有MD中的业务含义
        if table_name in existing_biz:
            fields = existing_biz[table_name].get('fields', {})
            if field_name in fields and fields[field_name]:
                return fields[field_name]
        # 2. 代码中的业务含义
        if table_name in code_biz:
            fields = code_biz[table_name].get('fields', {})
            if field_name in fields and fields[field_name]:
                return fields[field_name]
        # 3. 数据库注释
        if db_comment:
            return db_comment
        return ''

    def get_table_biz_meaning(table_name, table_comment=''):
        """获取表级业务含义"""
        if table_name in existing_biz:
            bm = existing_biz[table_name].get('table_biz_meaning', '')
            if bm:
                return bm
        if table_comment:
            return table_comment
        return ''

    # 4. 构建表结构
    tables_by_domain = defaultdict(list)

    for obj in objects:
        tname = obj['TABLE_NAME']
        ttype = obj['TABLE_TYPE']
        tcomment = obj['TABLE_COMMENT'] or ''
        engine = obj.get('ENGINE', '')
        rows = obj.get('TABLE_ROWS', 0) or 0
        data_len = obj.get('DATA_LENGTH', 0) or 0
        idx_len = obj.get('INDEX_LENGTH', 0) or 0

        domain = classify_table(tname, ttype)
        if domain is None:
            continue  # 过滤临时表

        # 字段列表
        fields = []
        for col in cols_by_table.get(tname, []):
            fname = col['COLUMN_NAME']
            db_comment = col.get('COLUMN_COMMENT', '') or ''
            biz = get_field_biz_meaning(tname, fname, db_comment)

            # 约束信息
            constraint_parts = []
            col_key = col.get('COLUMN_KEY', '')
            extra = col.get('EXTRA', '')
            if col_key == 'PRI':
                constraint_parts.append('PRI')
            elif col_key == 'UNI':
                constraint_parts.append('UNI')
            elif col_key == 'MUL':
                constraint_parts.append('MUL')
            if extra and 'auto_increment' in extra.lower():
                constraint_parts.append('AUTO_INCREMENT')

            # 默认值
            default_val = col.get('COLUMN_DEFAULT')
            if default_val is None:
                default_val = '-'
            else:
                default_val = str(default_val)

            fields.append({
                'name': fname,
                'type': col['COLUMN_TYPE'],
                'nullable': col['IS_NULLABLE'],
                'default': default_val,
                'constraint': ', '.join(constraint_parts),
                'comment': db_comment,
                'biz_meaning': biz,
            })

        # 约束列表
        table_constraints = constraints_by_table.get(tname, {})
        constraint_list = []
        for ctype, clist in table_constraints.items():
            for c in clist:
                constraint_list.append({
                    'name': c['CONSTRAINT_NAME'],
                    'type': ctype,
                    'column': c['COLUMN_NAME'],
                    'ref_table': c.get('REFERENCED_TABLE_NAME', ''),
                    'ref_column': c.get('REFERENCED_COLUMN_NAME', ''),
                })

        # 索引列表
        table_indexes = idx_by_table.get(tname, [])

        # 外键列表
        table_fks = fk_by_table.get(tname, [])

        # 数据大小
        total_size = data_len + idx_len
        if total_size > 1024 * 1024:
            size_str = f"{total_size / 1024 / 1024:.1f} MB"
        elif total_size > 1024:
            size_str = f"{total_size / 1024:.1f} KB"
        else:
            size_str = f"{total_size} B"

        table_info = {
            'table_name': tname,
            'table_type': ttype,
            'table_comment': tcomment,
            'biz_meaning': get_table_biz_meaning(tname, tcomment),
            'engine': engine,
            'rows': rows,
            'data_size': size_str,
            'fields': fields,
            'constraints': constraint_list,
            'indexes': table_indexes,
            'foreign_keys': table_fks,
        }

        tables_by_domain[domain].append(table_info)

    # 按表名排序
    for domain in tables_by_domain:
        tables_by_domain[domain].sort(key=lambda t: t['table_name'])

    return tables_by_domain


# ============================================================
# MD 生成
# ============================================================

def generate_md(tables_by_domain):
    lines = []
    lines.append('# dppms_d365 数据库完整数据字典')
    lines.append('')
    lines.append('> 生成时间：2026-06-13 | 数据库：dppms_d365 | 包含所有表和视图')
    lines.append('')
    lines.append('---')
    lines.append('')

    # 目录
    lines.append('## 目录')
    lines.append('')
    ch_idx = 0
    for domain in sorted(tables_by_domain.keys()):
        ch_idx += 1
        cn = ['一', '二', '三', '四', '五', '六', '七', '八'][ch_idx - 1] if ch_idx <= 8 else str(ch_idx)
        lines.append(f'- 第{cn}章 {domain} ({len(tables_by_domain[domain])} 个对象)')
    lines.append('')
    lines.append('---')
    lines.append('')

    # 各域
    ch_idx = 0
    for domain in sorted(tables_by_domain.keys()):
        ch_idx += 1
        cn = ['一', '二', '三', '四', '五', '六', '七', '八'][ch_idx - 1] if ch_idx <= 8 else str(ch_idx)
        tables = tables_by_domain[domain]
        lines.append(f'# 第{cn}章 {domain}')
        lines.append('')

        for tidx, t in enumerate(tables, 1):
            tname = t['table_name']
            tcomment = t['table_comment']
            title = f"### {ch_idx}.{tidx} {tname}"
            if tcomment:
                title += f" -- {tcomment}"
            lines.append(title)
            lines.append('')

            # 属性表
            lines.append('| 属性 | 值 |')
            lines.append('|------|-----|')
            lines.append(f"| 对象类型 | {t['table_type']} |")
            if t['engine']:
                lines.append(f"| 存储引擎 | {t['engine']} |")
            if t['biz_meaning']:
                lines.append(f"| 业务含义 | {t['biz_meaning']} |")
            if t['rows']:
                lines.append(f"| 数据量 | ~{t['rows']:,} 行 |")
            if t['data_size']:
                lines.append(f"| 数据大小 | {t['data_size']} |")
            lines.append('')

            # 字段列表
            if t['fields']:
                lines.append('**字段列表**')
                lines.append('')
                lines.append('| 字段名 | 数据类型 | 可空 | 默认值 | 约束 | 字段描述 | 业务含义 |')
                lines.append('|--------|----------|------|--------|------|----------|----------|')
                for f in t['fields']:
                    comment = f['comment'].replace('|', '\\|') if f['comment'] else ''
                    biz = f['biz_meaning'].replace('|', '\\|') if f['biz_meaning'] else ''
                    lines.append(f"| {f['name']} | {f['type']} | {f['nullable']} | {f['default']} | {f['constraint']} | {comment} | {biz} |")
                lines.append('')

            # 约束列表
            if t['constraints']:
                lines.append('**约束列表**')
                lines.append('')
                lines.append('| 约束名 | 约束类型 | 字段 | 引用表 | 引用字段 |')
                lines.append('|--------|----------|------|--------|----------|')
                for c in t['constraints']:
                    lines.append(f"| {c['name']} | {c['type']} | {c['column']} | {c['ref_table']} | {c['ref_column']} |")
                lines.append('')

            # 索引列表
            if t['indexes']:
                lines.append('**索引列表**')
                lines.append('')
                lines.append('| 索引名 | 索引类型 | 唯一性 | 索引字段 |')
                lines.append('|--------|----------|--------|----------|')
                for idx in t['indexes']:
                    uniqueness = 'UNIQUE' if not idx['NON_UNIQUE'] else 'NON-UNIQUE'
                    lines.append(f"| {idx['INDEX_NAME']} | {idx['INDEX_TYPE']} | {uniqueness} | {idx['COLUMNS']} |")
                lines.append('')

            # 外键列表
            if t['foreign_keys']:
                lines.append('**外键列表**')
                lines.append('')
                lines.append('| 外键名 | 本表字段 | 引用表 | 引用字段 |')
                lines.append('|--------|----------|--------|----------|')
                for fk in t['foreign_keys']:
                    lines.append(f"| {fk['CONSTRAINT_NAME']} | {fk['COLUMN_NAME']} | {fk['REFERENCED_TABLE_NAME']} | {fk['REFERENCED_COLUMN_NAME']} |")
                lines.append('')

            lines.append('---')
            lines.append('')

    with open(MD_OUT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))
    print(f"MD 已保存: {MD_OUT} ({len(lines)} 行)")


# ============================================================
# Excel 生成
# ============================================================

# 样式定义
COLOR_TITLE_BG = '1F4E79'
COLOR_TITLE_FONT = 'FFFFFF'
COLOR_ATTR_KEY_BG = 'B4C6E7'
COLOR_ATTR_VAL_BG = 'D6E4F0'
COLOR_FIELD_HDR_BG = '4472C4'
COLOR_FIELD_HDR_FONT = 'FFFFFF'
COLOR_CONSTRAINT_HDR_BG = '7030A0'
COLOR_CONSTRAINT_HDR_FONT = 'FFFFFF'
COLOR_INDEX_HDR_BG = '548235'
COLOR_INDEX_HDR_FONT = 'FFFFFF'
COLOR_FK_HDR_BG = 'BF8F00'
COLOR_FK_HDR_FONT = 'FFFFFF'
COLOR_ALT_ROW = 'F2F7FB'
COLOR_BORDER = 'B4C6E7'

FONT_TITLE = Font(name='微软雅黑', bold=True, size=12, color=COLOR_TITLE_FONT)
FONT_ATTR_KEY = Font(name='微软雅黑', bold=True, size=10, color='1F4E79')
FONT_ATTR_VAL = Font(name='微软雅黑', size=10, color='333333')
FONT_FIELD_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_FIELD_HDR_FONT)
FONT_FIELD_DATA = Font(name='微软雅黑', size=9, color='333333')
FONT_CONSTRAINT_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_CONSTRAINT_HDR_FONT)
FONT_CONSTRAINT_DATA = Font(name='微软雅黑', size=9, color='333333')
FONT_INDEX_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_INDEX_HDR_FONT)
FONT_INDEX_DATA = Font(name='微软雅黑', size=9, color='333333')
FONT_FK_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_FK_HDR_FONT)
FONT_FK_DATA = Font(name='微软雅黑', size=9, color='333333')
FONT_SECTION = Font(name='微软雅黑', bold=True, size=11, color='2E75B6')

FILL_TITLE = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type='solid')
FILL_ATTR_KEY = PatternFill(start_color=COLOR_ATTR_KEY_BG, end_color=COLOR_ATTR_KEY_BG, fill_type='solid')
FILL_ATTR_VAL = PatternFill(start_color=COLOR_ATTR_VAL_BG, end_color=COLOR_ATTR_VAL_BG, fill_type='solid')
FILL_FIELD_HDR = PatternFill(start_color=COLOR_FIELD_HDR_BG, end_color=COLOR_FIELD_HDR_BG, fill_type='solid')
FILL_CONSTRAINT_HDR = PatternFill(start_color=COLOR_CONSTRAINT_HDR_BG, end_color=COLOR_CONSTRAINT_HDR_BG, fill_type='solid')
FILL_INDEX_HDR = PatternFill(start_color=COLOR_INDEX_HDR_BG, end_color=COLOR_INDEX_HDR_BG, fill_type='solid')
FILL_FK_HDR = PatternFill(start_color=COLOR_FK_HDR_BG, end_color=COLOR_FK_HDR_BG, fill_type='solid')
FILL_ALT_ROW = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid')

ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_TITLE = Alignment(horizontal='left', vertical='center')

THIN_SIDE = Side(style='thin', color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

MAX_COL = 9  # A-I

COL_WIDTHS_FIELD = {'A': 28, 'B': 20, 'C': 8, 'D': 16, 'E': 16, 'F': 30, 'G': 40, 'H': 16, 'I': 16}


def write_table_excel(ws, table, start_row):
    row = start_row

    # 1. 表名标题行
    title_text = table['table_name']
    if table['table_comment']:
        title_text += f"  ——  {table['table_comment']}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
    cell = ws.cell(row=row, column=1, value=title_text)
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    cell.alignment = ALIGN_TITLE
    cell.border = BORDER_ALL
    for c in range(2, MAX_COL + 1):
        ws.cell(row=row, column=c).fill = FILL_TITLE
        ws.cell(row=row, column=c).border = BORDER_ALL
    row += 1

    # 2. 属性区
    attrs = [
        ('对象类型', table['table_type']),
        ('业务含义', table['biz_meaning']),
        ('数据量', f"~{table['rows']:,} 行" if table['rows'] else ''),
        ('数据大小', table['data_size']),
    ]
    for key, val in attrs:
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=MAX_COL)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell_key = ws.cell(row=row, column=1, value=key)
        cell_key.font = FONT_ATTR_KEY
        cell_key.fill = FILL_ATTR_KEY
        cell_key.alignment = ALIGN_CENTER
        cell_key.border = BORDER_ALL
        ws.cell(row=row, column=2).fill = FILL_ATTR_KEY
        ws.cell(row=row, column=2).border = BORDER_ALL
        cell_val = ws.cell(row=row, column=3, value=val)
        cell_val.font = FONT_ATTR_VAL
        cell_val.fill = FILL_ATTR_VAL
        cell_val.alignment = ALIGN_LEFT
        cell_val.border = BORDER_ALL
        for c in range(4, MAX_COL + 1):
            ws.cell(row=row, column=c).fill = FILL_ATTR_VAL
            ws.cell(row=row, column=c).border = BORDER_ALL
        row += 1

    # 3. 字段列表
    field_headers = ['字段名', '数据类型', '可空', '默认值', '约束', '字段描述', '业务含义', '', '']
    for col_idx, hdr in enumerate(field_headers[:7], 1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = FONT_FIELD_HDR
        cell.fill = FILL_FIELD_HDR
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    for c in range(8, MAX_COL + 1):
        ws.cell(row=row, column=c).fill = FILL_FIELD_HDR
        ws.cell(row=row, column=c).border = BORDER_ALL
    row += 1

    for f_idx, f in enumerate(table['fields']):
        is_alt = f_idx % 2 == 1
        row_data = [f['name'], f['type'], f['nullable'], f['default'], f['constraint'], f['comment'], f['biz_meaning']]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_FIELD_DATA
            cell.alignment = ALIGN_LEFT if col_idx in (1, 6, 7) else ALIGN_CENTER
            cell.border = BORDER_ALL
            if is_alt:
                cell.fill = FILL_ALT_ROW
        row += 1

    # 4. 约束列表
    if table['constraints']:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        cell = ws.cell(row=row, column=1, value='约束列表')
        cell.font = FONT_SECTION
        cell.alignment = ALIGN_LEFT
        row += 1

        constraint_headers = ['约束名', '约束类型', '字段', '引用表', '引用字段']
        for col_idx, hdr in enumerate(constraint_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=hdr)
            cell.font = FONT_CONSTRAINT_HDR
            cell.fill = FILL_CONSTRAINT_HDR
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        row += 1

        for c_idx, c in enumerate(table['constraints']):
            is_alt = c_idx % 2 == 1
            row_data = [c['name'], c['type'], c['column'], c['ref_table'], c['ref_column']]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = FONT_CONSTRAINT_DATA
                cell.alignment = ALIGN_LEFT
                cell.border = BORDER_ALL
                if is_alt:
                    cell.fill = FILL_ALT_ROW
            row += 1

    # 5. 索引列表
    if table['indexes']:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        cell = ws.cell(row=row, column=1, value='索引列表')
        cell.font = FONT_SECTION
        cell.alignment = ALIGN_LEFT
        row += 1

        index_headers = ['索引名', '索引类型', '唯一性', '索引字段']
        for col_idx, hdr in enumerate(index_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=hdr)
            cell.font = FONT_INDEX_HDR
            cell.fill = FILL_INDEX_HDR
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=MAX_COL)
        row += 1

        for i_idx, idx in enumerate(table['indexes']):
            is_alt = i_idx % 2 == 1
            uniqueness = 'UNIQUE' if not idx['NON_UNIQUE'] else 'NON-UNIQUE'
            row_data = [idx['INDEX_NAME'], idx['INDEX_TYPE'], uniqueness, idx['COLUMNS']]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = FONT_INDEX_DATA
                cell.alignment = ALIGN_LEFT if col_idx == 4 else ALIGN_CENTER
                cell.border = BORDER_ALL
                if is_alt:
                    cell.fill = FILL_ALT_ROW
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=MAX_COL)
            row += 1

    # 6. 外键列表
    if table['foreign_keys']:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COL)
        cell = ws.cell(row=row, column=1, value='外键列表')
        cell.font = FONT_SECTION
        cell.alignment = ALIGN_LEFT
        row += 1

        fk_headers = ['外键名', '本表字段', '引用表', '引用字段']
        for col_idx, hdr in enumerate(fk_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=hdr)
            cell.font = FONT_FK_HDR
            cell.fill = FILL_FK_HDR
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        row += 1

        for fk_idx, fk in enumerate(table['foreign_keys']):
            is_alt = fk_idx % 2 == 1
            row_data = [fk['CONSTRAINT_NAME'], fk['COLUMN_NAME'], fk['REFERENCED_TABLE_NAME'], fk['REFERENCED_COLUMN_NAME']]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = FONT_FK_DATA
                cell.alignment = ALIGN_LEFT
                cell.border = BORDER_ALL
                if is_alt:
                    cell.fill = FILL_ALT_ROW
            row += 1

    row += 2  # 表间间隔
    return row


def generate_xlsx(tables_by_domain):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for domain in sorted(tables_by_domain.keys()):
        tables = tables_by_domain[domain]
        if not tables:
            continue

        sheet_name = domain[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 设置列宽
        for col_letter, width in COL_WIDTHS_FIELD.items():
            ws.column_dimensions[col_letter].width = width

        row = 1
        for table in tables:
            row = write_table_excel(ws, table, row)

        ws.freeze_panes = 'A2'
        print(f"   {domain}: {len(tables)} 个对象, {row} 行")

    wb.save(XLSX_OUT)
    print(f"Excel 已保存: {XLSX_OUT}")


# ============================================================
# 主流程
# ============================================================

def main():
    print("=" * 70)
    print("生成完整数据字典（MD + Excel）")
    print("数据源：数据库schema + 现有MD业务含义 + 代码业务含义")
    print("=" * 70)

    # 1. 构建数据
    print("\n1. 合并数据源...")
    tables_by_domain = build_data()

    # 2. 统计
    print("\n2. 统计:")
    total_tables = 0
    total_views = 0
    total_fields = 0
    total_constraints = 0
    total_indexes = 0
    total_fks = 0
    total_biz_meaning = 0

    for domain, tables in sorted(tables_by_domain.items()):
        t_count = sum(1 for t in tables if t['table_type'] == 'BASE TABLE')
        v_count = sum(1 for t in tables if t['table_type'] == 'VIEW')
        f_count = sum(len(t['fields']) for t in tables)
        c_count = sum(len(t['constraints']) for t in tables)
        i_count = sum(len(t['indexes']) for t in tables)
        fk_count = sum(len(t['foreign_keys']) for t in tables)
        bm_count = sum(1 for t in tables for f in t['fields'] if f['biz_meaning'])

        total_tables += t_count
        total_views += v_count
        total_fields += f_count
        total_constraints += c_count
        total_indexes += i_count
        total_fks += fk_count
        total_biz_meaning += bm_count

        print(f"   {domain}: {t_count} 表 + {v_count} 视图, {f_count} 字段, {c_count} 约束, {i_count} 索引, {fk_count} 外键")

    print(f"\n   总计: {total_tables} 表 + {total_views} 视图")
    print(f"   总字段数: {total_fields}")
    print(f"   有业务含义的字段: {total_biz_meaning} ({total_biz_meaning*100//total_fields if total_fields else 0}%)")
    print(f"   总约束数: {total_constraints}")
    print(f"   总索引数: {total_indexes}")
    print(f"   总外键数: {total_fks}")

    # 3. 生成MD
    print("\n3. 生成 MD...")
    generate_md(tables_by_domain)

    # 4. 生成Excel
    print("\n4. 生成 Excel...")
    generate_xlsx(tables_by_domain)

    print("\n完成!")


if __name__ == '__main__':
    main()
