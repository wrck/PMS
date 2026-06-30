#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证完整数据字典Excel"""
from openpyxl import load_workbook

wb = load_workbook('d:\\EclipseWorkspace\\Parctice\\PMS\\PMS-struts\\docs\\03-database\\database_dict_full.xlsx')
print("Sheet 列表:")
total_tables = 0
total_views = 0
empty_fields = 0
for name in wb.sheetnames:
    ws = wb[name]
    t_count = 0
    v_count = 0
    e_count = 0
    row = 1
    while row <= ws.max_row:
        cell = ws.cell(row=row, column=1)
        if cell.font and cell.font.bold and cell.font.size == 12 and cell.font.color and cell.font.color.rgb and 'FFFFFF' in str(cell.font.color.rgb):
            title = str(cell.value or '')
            # 检查是否视图
            obj_type_row = row + 1
            if obj_type_row <= ws.max_row:
                ot = ws.cell(row=obj_type_row, column=3)
                if ot.value and 'VIEW' in str(ot.value):
                    v_count += 1
                else:
                    t_count += 1
            else:
                t_count += 1
            # 检查字段数据
            field_header_row = row + 5
            has_data = False
            if field_header_row <= ws.max_row and ws.cell(row=field_header_row, column=1).value == '字段名':
                check_row = field_header_row + 1
                if check_row <= ws.max_row:
                    dc = ws.cell(row=check_row, column=1)
                    if dc.value and str(dc.value).strip():
                        has_data = True
            if not has_data:
                e_count += 1
                print(f'  [空字段] {title[:50]}')
            row += 6
        else:
            row += 1
    total_tables += t_count
    total_views += v_count
    empty_fields += e_count
    print(f"  {name}: {t_count} 表 + {v_count} 视图, 空字段: {e_count}")

print(f"\n总计: {total_tables} 表 + {total_views} 视图, 空字段表: {empty_fields}")
