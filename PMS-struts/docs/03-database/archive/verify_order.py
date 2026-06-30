#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证合并后pm_project表的字段顺序（包括无字段名的新增行）"""
import openpyxl

wb = openpyxl.load_workbook(r'd:\EclipseWorkspace\Parctice\PMS\PMS-struts\docs\03-database\database_dict_flat_final_merged_v2.xlsx')
ws = wb['项目管理']

in_table = False
count = 0
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    c = ws.cell(r, 3).value
    h = ws.cell(r, 8).value  # 字段描述

    if a and str(a).strip() == 'pm_project':
        in_table = True
    elif a and str(a).strip() != 'pm_project' and in_table and str(a).strip() != '表名':
        break

    if in_table and (c or h):
        fname = str(c).strip() if c else ''
        comment = str(h or '').strip()
        biz_match = str(ws.cell(r, 11).value or '').strip()  # 对应业务

        is_new = "★新增" if (not fname and comment) else ""
        print(f"  {count+1:2d}. fname='{fname:25s}' comment='{comment[:25]:25s}' 对应业务={biz_match:15s} {is_new}")
        count += 1

print(f"\npm_project 合并后总字段数: {count}")
