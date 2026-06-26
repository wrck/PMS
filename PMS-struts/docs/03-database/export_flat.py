#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
新格式导出：将 database_dict final.md 导出为 Excel
- 无表头属性区
- 表名和描述作为字段列表的前2列
- 发货相关表移至项目管理域
- 按章节顺序输出
"""

import re
import os
import copy
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
MD_FILE = os.path.join(BASE, 'database_dict final.md')
XLSX_FILE = os.path.join(BASE, 'database_dict_flat_final.xlsx')

# ============================================================
# 业务域定义
# ============================================================
DOMAINS = [
    ("项目管理", "第一章", ["一、项目管理域", "二、回访管理域", "三、售前管理域",
                          "四、转包管理域", "五、问题管理域", "六、基础平台域",
                          "七、维保管理域", "八、安全行业资产管理域", "九、项目辅助表"]),
    ("系统支撑", "第二章", ["一、EHR组织架构域", "二、系统权限域",
                          "三、数据同步中间表域", "四、其他辅助表"]),
    ("历史迁移与引擎", "第三章", ["一、Activiti工作流引擎表", "二、Firebird迁移表",
                                "三、RMA/备件/仓库等业务表"]),
]

# 发货相关的表（从历史迁移域移至项目管理域）
SHIPMENT_TABLES = {
    'fb_shipment', 'fb_shipment_barcode', 'fb_shipment_barcode_change_log',
    'fb_shipment_barcode_order_line', 'fb_shipment_barcode_relation',
    'fb_items', 'fb_items2',
    'shipment_barcode_from_spms_unique',
}

# ============================================================
# 过滤规则
# ============================================================
def should_filter(table_name, obj_type):
    if table_name.startswith('temp_') or table_name.startswith('tmp_'):
        return True
    if obj_type == 'VIEW':
        return True
    return False

# ============================================================
# 样式定义
# ============================================================
COLOR_TITLE_BG = '1F4E79'
COLOR_TITLE_FONT = 'FFFFFF'
COLOR_HDR_BG = '2E75B6'
COLOR_HDR_FONT = 'FFFFFF'
COLOR_ALT_ROW = 'F2F7FB'
COLOR_BORDER = 'B4C6E7'

FONT_TITLE = Font(name='微软雅黑', bold=True, size=11, color=COLOR_TITLE_FONT)
FONT_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_HDR_FONT)
FONT_DATA = Font(name='微软雅黑', size=9, color='333333')

FILL_TITLE = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type='solid')
FILL_HDR = PatternFill(start_color=COLOR_HDR_BG, end_color=COLOR_HDR_BG, fill_type='solid')
FILL_ALT = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid')

ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN_SIDE = Side(style='thin', color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# 列宽
COL_WIDTHS = {
    'A': 30,  # 表名
    'B': 28,  # 表描述
    'C': 22,  # 字段名
    'D': 18,  # 数据类型
    'E': 8,   # 可空
    'F': 14,  # 默认值
    'G': 14,  # 约束
    'H': 32,  # 字段描述
    'I': 40,  # 业务含义
}

# 字段列表表头
HEADERS = ['表名', '表描述', '字段名', '数据类型', '可空', '默认值', '约束', '字段描述', '业务含义']

# ============================================================
# 解析逻辑（与 export_dict_to_excel.py 相同）
# ============================================================

def read_md(filepath):
    with open(filepath, 'rb') as f:
        data = f.read()
    text = data.decode('utf-8', errors='replace')
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    return text


def find_chapter_boundaries(lines):
    chapters = {}
    ch1 = ch2 = ch3 = None
    for i, line in enumerate(lines):
        s = line.strip()
        if s.startswith('# 第一章'):
            ch1 = i
        elif s.startswith('# 第二章'):
            ch2 = i
        elif s.startswith('# 第三章'):
            ch3 = i
    if ch1 is not None:
        chapters['第一章'] = (ch1, ch2 if ch2 else len(lines))
    if ch2 is not None:
        chapters['第二章'] = (ch2, ch3 if ch3 else len(lines))
    if ch3 is not None:
        chapters['第三章'] = (ch3, len(lines))
    return chapters


def find_subdomain_boundaries(lines, start, end):
    subdomains = []
    for i in range(start, end):
        s = lines[i].strip()
        m = re.match(r'^#{1,2}\s+([一二三四五六七八九十]+、.+)', s)
        if m:
            subdomains.append((m.group(1).strip(), i))
    result = []
    for idx, (name, line_no) in enumerate(subdomains):
        next_line = subdomains[idx + 1][1] if idx + 1 < len(subdomains) else end
        result.append((name, line_no, next_line))
    return result


def parse_field_row(raw_line):
    all_parts = raw_line.split('|')
    if all_parts and all_parts[0].strip() == '':
        all_parts = all_parts[1:]
    if all_parts and all_parts[-1].strip() == '':
        all_parts = all_parts[:-1]
    stripped = [p.strip() for p in all_parts]
    col_count = len(stripped)

    if col_count >= 7:
        return {
            '字段名': stripped[0], '数据类型': stripped[1], '可空': stripped[2],
            '默认值': stripped[3], '约束': stripped[4],
            '字段描述': ' | '.join(stripped[5:-1]), '业务含义': stripped[-1],
        }
    elif col_count == 6:
        fifth = stripped[4]
        if fifth in ('PRI', 'UNI', 'MUL', 'AUTO_INCREMENT') or 'auto_increment' in fifth.lower() or fifth in ('-', ''):
            return {
                '字段名': stripped[0], '数据类型': stripped[1], '可空': stripped[2],
                '默认值': stripped[3], '约束': fifth,
                '字段描述': stripped[5], '业务含义': '',
            }
        else:
            return {
                '字段名': stripped[0], '数据类型': stripped[1], '可空': stripped[2],
                '默认值': stripped[3], '约束': '',
                '字段描述': ' | '.join(stripped[4:-1]),
                '业务含义': stripped[-1] if stripped[-1] != stripped[4] else '',
            }
    elif col_count == 5:
        return {
            '字段名': stripped[0], '数据类型': stripped[1], '可空': stripped[2],
            '默认值': stripped[3], '约束': '',
            '字段描述': stripped[4], '业务含义': '',
        }
    return None


def parse_tables(lines, start, end):
    tables = []
    i = start
    while i < end:
        line = lines[i].strip()
        m_std = re.match(r'^#{3,5}\s+[\d.]*\s*(\S+)\s*(?:--\s*(.+))?', line)
        if not m_std or not is_valid_table_name(m_std.group(1)):
            i += 1
            continue

        table_name = m_std.group(1)
        table_comment = m_std.group(2) or ''
        i += 1

        obj_type, biz_meaning, data_count, data_size = '', '', '', ''
        while i < end:
            l = lines[i].strip()
            if l.startswith('|') and '属性' in l and '值' in l:
                i += 1; i += 1
                while i < end and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    parts = [p.strip() for p in lines[i].strip().split('|') if p.strip()]
                    if len(parts) >= 2:
                        key, val = parts[0], parts[1]
                        if '对象类型' in key: obj_type = val
                        elif '业务含义' in key: biz_meaning = val
                        elif '数据量' in key: data_count = val
                        elif '数据大小' in key or '存储大小' in key: data_size = val
                    i += 1
                break
            elif l.startswith('**字段列表**') or l.startswith('| 字段名'):
                break
            else:
                i += 1

        if should_filter(table_name, obj_type):
            while i < end:
                l = lines[i].strip()
                if re.match(r'^#{2,5}\s+[\d.]*\s*\S+', l) or l == '---':
                    break
                i += 1
            continue

        fields = []
        ref_table = ''
        while i < end:
            l = lines[i].strip()
            if l.startswith('| 字段名'):
                i += 1; i += 1
                while i < end and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    field = parse_field_row(lines[i].strip())
                    if field:
                        fields.append(field)
                    i += 1
                break
            elif l.startswith('**索引') or l.startswith('**外键') or l.startswith('---') or re.match(r'^#{2,5}', l):
                break
            else:
                m_ref = re.match(r'^与\s+(\S+)\s+结构(相同|类似|基本相同)', l)
                if m_ref:
                    ref_table = m_ref.group(1)
                i += 1

        indexes = []
        while i < end:
            l = lines[i].strip()
            if l.startswith('| 索引名') and '索引类型' in l:
                i += 1; i += 1
                while i < end and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    non_empty = [p.strip() for p in lines[i].strip().split('|') if p.strip()]
                    if len(non_empty) >= 4:
                        indexes.append({
                            '索引名': non_empty[0], '索引类型': non_empty[1],
                            '唯一性': non_empty[2], '索引字段': ' | '.join(non_empty[3:]),
                        })
                    i += 1
                break
            elif l.startswith('**外键') or l.startswith('---') or re.match(r'^#{2,5}', l):
                break
            else:
                i += 1

        foreign_keys = []
        while i < end:
            l = lines[i].strip()
            if l.startswith('| 外键名') and '本表字段' in l:
                i += 1; i += 1
                while i < end and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    non_empty = [p.strip() for p in lines[i].strip().split('|') if p.strip()]
                    if len(non_empty) >= 4:
                        foreign_keys.append({
                            '外键名': non_empty[0], '本表字段': non_empty[1],
                            '引用表': non_empty[2], '引用字段': non_empty[3],
                        })
                    i += 1
                break
            elif l.startswith('---') or re.match(r'^#{2,5}', l):
                break
            else:
                i += 1

        tables.append({
            'table_name': table_name,
            'table_comment': table_comment,
            'obj_type': obj_type or 'BASE TABLE',
            'biz_meaning': biz_meaning,
            'data_count': data_count,
            'data_size': data_size,
            'fields': fields,
            'indexes': indexes,
            'foreign_keys': foreign_keys,
            'ref_table': ref_table,
        })

        while i < end:
            l = lines[i].strip()
            if l == '---' or re.match(r'^#{2,5}\s+', l):
                break
            i += 1

    return tables


def is_valid_table_name(name):
    if not name or len(name) < 2:
        return False
    if re.search(r'[\u4e00-\u9fff]', name):
        return False
    if name.startswith('**'):
        return False
    return True


# ============================================================
# Excel 写入逻辑（新格式：扁平化，表名+描述作为前2列）
# ============================================================

def write_tables_to_sheet(ws, tables):
    """将多个表写入工作表：表名+表描述作为前2列，每个字段行都填充，表间空1行"""
    row = 1
    num_cols = len(HEADERS)

    for table in tables:
        tname = table['table_name']
        tcomment = table['table_comment'] or table['biz_meaning'] or ''

        # 列名表头行
        for col_idx, hdr in enumerate(HEADERS, 1):
            cell = ws.cell(row=row, column=col_idx, value=hdr)
            cell.font = FONT_HDR
            cell.fill = FILL_HDR
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        row += 1

        # 字段数据行（表名+表描述每行都填充）
        for f_idx, field in enumerate(table['fields']):
            is_alt = f_idx % 2 == 1
            row_data = [
                tname,
                tcomment,
                field.get('字段名', ''),
                field.get('数据类型', ''),
                field.get('可空', ''),
                field.get('默认值', ''),
                field.get('约束', ''),
                field.get('字段描述', ''),
                field.get('业务含义', ''),
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = FONT_DATA
                cell.alignment = ALIGN_LEFT if col_idx in (1, 2, 3, 8, 9) else ALIGN_CENTER
                cell.border = BORDER_ALL
                if is_alt:
                    cell.fill = FILL_ALT
            row += 1

        # 空行（表间分隔）
        row += 1

    return row


def main():
    print("=" * 60)
    print("导出 database_dict_flat.xlsx（扁平格式）")
    print("格式: 表名+描述作为前2列，无表头属性区")
    print("过滤: temp_*/tmp_* 表 + 视图")
    print("=" * 60)

    # 1. 读取MD
    print("\n1. 读取 MD 文件...")
    text = read_md(MD_FILE)
    lines = text.split('\n')
    print(f"   总行数: {len(lines)}")

    # 2. 识别章节
    print("\n2. 识别章节边界...")
    chapters = find_chapter_boundaries(lines)

    # 3. 解析各业务域
    print("\n3. 解析各业务域表结构...")
    tables_by_domain = {}
    all_tables_global = {}

    for domain_name, chapter_key, subdomain_keywords in DOMAINS:
        if chapter_key not in chapters:
            continue
        ch_start, ch_end = chapters[chapter_key]
        subdomains = find_subdomain_boundaries(lines, ch_start, ch_end)

        target_subdomains = []
        for sd_name, sd_start, sd_end in subdomains:
            for kw in subdomain_keywords:
                if kw in sd_name or sd_name.startswith(kw):
                    target_subdomains.append((sd_name, sd_start, sd_end))
                    break

        if not target_subdomains:
            target_subdomains = [(chapter_key, ch_start, ch_end)]

        all_tables = []
        for sd_name, sd_start, sd_end in target_subdomains:
            tables = parse_tables(lines, sd_start, sd_end)
            all_tables.extend(tables)

        tables_by_domain[domain_name] = all_tables
        for t in all_tables:
            all_tables_global[t['table_name']] = t
        print(f"   {domain_name}: {len(all_tables)} 张表")

    # 4. 将发货相关表从历史迁移域移至项目管理域
    print("\n4. 移动发货相关表...")
    moved_count = 0
    shipment_tables = []
    remaining = []
    for t in tables_by_domain.get('历史迁移与引擎', []):
        if t['table_name'] in SHIPMENT_TABLES:
            shipment_tables.append(t)
            moved_count += 1
            print(f"   [移动] {t['table_name']} -> 项目管理域")
        else:
            remaining.append(t)
    tables_by_domain['历史迁移与引擎'] = remaining
    # 插入到项目管理域末尾
    tables_by_domain['项目管理'].extend(shipment_tables)
    print(f"   共移动 {moved_count} 张表")

    # 5. 解析引用
    print("\n5. 解析结构引用...")
    ref_resolved = 0
    for domain_name, tables in tables_by_domain.items():
        for t in tables:
            if not t['fields'] and t.get('ref_table'):
                ref_name = t['ref_table']
                if ref_name in all_tables_global and all_tables_global[ref_name]['fields']:
                    t['fields'] = copy.deepcopy(all_tables_global[ref_name]['fields'])
                    t['indexes'] = copy.deepcopy(all_tables_global[ref_name]['indexes'])
                    ref_resolved += 1
                    print(f"   [引用解析] {t['table_name']} <- {ref_name}")

    # 6. 查库补全
    empty_tables = [(dn, t) for dn, tables in tables_by_domain.items() for t in tables if not t['fields']]
    if empty_tables:
        print(f"\n6. 查库补全 {len(empty_tables)} 张空字段表...")
        try:
            conn = pymysql.connect(host='localhost', user='root', password='!Q@W3e4r', database='dppms_d365', charset='utf8mb4')
            cursor = conn.cursor()
            for dn, t in empty_tables:
                cursor.execute(
                    "SELECT COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT, COLUMN_KEY, EXTRA, COLUMN_COMMENT "
                    "FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA='dppms_d365' AND TABLE_NAME=%s ORDER BY ORDINAL_POSITION",
                    (t['table_name'],)
                )
                rows = cursor.fetchall()
                for row in rows:
                    col_name, col_type, is_nullable, col_default, col_key, extra, col_comment = row
                    constraint = ''
                    if col_key == 'PRI': constraint = 'PRI'
                    elif col_key == 'UNI': constraint = 'UNI'
                    elif col_key == 'MUL': constraint = 'MUL'
                    if extra and 'auto_increment' in extra:
                        constraint += ', AUTO_INCREMENT' if constraint else 'AUTO_INCREMENT'
                    default_val = str(col_default) if col_default is not None else '-'
                    if default_val == 'None': default_val = '-'
                    t['fields'].append({
                        '字段名': col_name, '数据类型': col_type, '可空': is_nullable,
                        '默认值': default_val, '约束': constraint,
                        '字段描述': col_comment or '', '业务含义': '',
                    })
                print(f"   [补全] {t['table_name']}: {len(rows)} 个字段")
            cursor.close(); conn.close()
        except Exception as e:
            print(f"   [数据库连接失败] {e}")
    else:
        print("\n6. 所有表均有字段数据，无需查库补全")

    # 7. 写入 Excel
    print("\n7. 写入 Excel...")
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for domain_name, tables in tables_by_domain.items():
        if not tables:
            continue
        sheet_name = domain_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        for col_letter, width in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        row = write_tables_to_sheet(ws, tables)
        ws.freeze_panes = 'A2'

        total_fields = sum(len(t['fields']) for t in tables)
        print(f"   {domain_name}: {len(tables)} 张表, {total_fields} 字段, {row} 行")

    wb.save(XLSX_FILE)
    print(f"\n8. Excel 已保存: {XLSX_FILE}")

    # 统计
    total_tables = sum(len(t) for t in tables_by_domain.values())
    total_fields = sum(len(t['fields']) for tables in tables_by_domain.values() for t in tables)
    has_comment = sum(1 for tables in tables_by_domain.values() for t in tables for f in t['fields'] if f.get('字段描述') and f['字段描述'] not in ('-', 'None', ''))
    has_biz = sum(1 for tables in tables_by_domain.values() for t in tables for f in t['fields'] if f.get('业务含义') and f['业务含义'] not in ('-', 'None', '', '业务含义待确认'))

    print(f"\n统计:")
    print(f"  总表数: {total_tables}")
    print(f"  总字段数: {total_fields}")
    print(f"  有字段描述: {has_comment} ({has_comment*100//total_fields}%)")
    print(f"  有业务含义: {has_biz} ({has_biz*100//total_fields}%)")


if __name__ == '__main__':
    main()
