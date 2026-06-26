#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""详细检查匹配表的结构"""
import openpyxl

wb2 = openpyxl.load_workbook(r'd:\EclipseWorkspace\Parctice\PMS\PMS-struts\docs\03-database\项目管理表数据元.xlsx')
ws = wb2[wb2.sheetnames[0]]

print(f"Sheet: {wb2.sheetnames[0]}, rows={ws.max_row}, cols={ws.max_column}")

# 打印前30行，看清楚结构
for r in range(1, min(35, ws.max_row+1)):
    row_data = []
    for c in range(1, min(20, ws.max_column+1)):
        v = ws.cell(r, c).value
        if v is not None:
            row_data.append(f"C{c}={str(v)[:25]}")
    if row_data:
        print(f"  Row {r}: {', '.join(row_data)}")
    else:
        print(f"  Row {r}: [empty]")
