#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
将 database_dict.md 解析并导出为专业格式 Excel 文件
- 按业务域分 sheet
- 过滤 temp_*/tmp_* 表和视图
- 每个表独立呈现（表头属性 + 字段列表 + 索引列表）
- 专业格式：字体、对齐、边框、颜色
"""

import re
import os
import copy
import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

MD_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database_dict final.md')
XLSX_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'database_dict_final_v2.xlsx')

# ============================================================
# 业务域定义（sheet名, 章节关键词, 子域关键词列表）
# ============================================================
DOMAINS = [
    ("项目管理", "第一章", ["一、项目管理域", "二、回访管理域", "三、售前管理域",
                          "四、转包管理域", "五、问题管理域", "六、基础平台域",
                          "七、维保管理域", "八、安全行业资产管理域", "九、项目辅助表"]),
    ("系统支撑", "第二章", ["一、EHR组织架构域", "二、系统权限域",
                          "三、数据同步中间表域", "四、其他辅助表"]),
    ("历史迁移与引擎", "第三章", ["一、Activiti工作流引擎表", "二、Firebird迁移表",
                                "三、RMA/备件/仓库等业务表"]),
    # 注意：不包含"四、临时表"和"五、视图"
]

# ============================================================
# 过滤规则
# ============================================================
def should_filter(table_name, obj_type):
    """过滤 temp_*/tmp_* 表和视图"""
    if table_name.startswith('temp_') or table_name.startswith('tmp_'):
        return True
    if obj_type == 'VIEW':
        return True
    return False

# ============================================================
# 样式定义
# ============================================================
# 颜色
COLOR_TITLE_BG = '1F4E79'      # 深蓝 - 表名标题
COLOR_TITLE_FONT = 'FFFFFF'    # 白色 - 表名标题字体
COLOR_ATTR_BG = 'D6E4F0'       # 浅蓝 - 属性区
COLOR_ATTR_KEY_BG = 'B4C6E7'   # 中蓝 - 属性键
COLOR_FIELD_HDR_BG = '4472C4'  # 蓝色 - 字段表头
COLOR_FIELD_HDR_FONT = 'FFFFFF'
COLOR_INDEX_HDR_BG = '548235'  # 绿色 - 索引表头
COLOR_INDEX_HDR_FONT = 'FFFFFF'
COLOR_FK_HDR_BG = 'BF8F00'    # 金色 - 外键表头
COLOR_FK_HDR_FONT = 'FFFFFF'
COLOR_ALT_ROW = 'F2F7FB'       # 交替行浅色
COLOR_BORDER = 'B4C6E7'       # 边框色

# 字体
FONT_TITLE = Font(name='微软雅黑', bold=True, size=12, color=COLOR_TITLE_FONT)
FONT_ATTR_KEY = Font(name='微软雅黑', bold=True, size=10, color='1F4E79')
FONT_ATTR_VAL = Font(name='微软雅黑', size=10, color='333333')
FONT_FIELD_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_FIELD_HDR_FONT)
FONT_FIELD_DATA = Font(name='微软雅黑', size=9, color='333333')
FONT_INDEX_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_INDEX_HDR_FONT)
FONT_INDEX_DATA = Font(name='微软雅黑', size=9, color='333333')
FONT_FK_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_FK_HDR_FONT)
FONT_FK_DATA = Font(name='微软雅黑', size=9, color='333333')
FONT_SECTION = Font(name='微软雅黑', bold=True, size=11, color='2E75B6')

# 填充
FILL_TITLE = PatternFill(start_color=COLOR_TITLE_BG, end_color=COLOR_TITLE_BG, fill_type='solid')
FILL_ATTR_KEY = PatternFill(start_color=COLOR_ATTR_KEY_BG, end_color=COLOR_ATTR_KEY_BG, fill_type='solid')
FILL_ATTR_VAL = PatternFill(start_color=COLOR_ATTR_BG, end_color=COLOR_ATTR_BG, fill_type='solid')
FILL_FIELD_HDR = PatternFill(start_color=COLOR_FIELD_HDR_BG, end_color=COLOR_FIELD_HDR_BG, fill_type='solid')
FILL_INDEX_HDR = PatternFill(start_color=COLOR_INDEX_HDR_BG, end_color=COLOR_INDEX_HDR_BG, fill_type='solid')
FILL_FK_HDR = PatternFill(start_color=COLOR_FK_HDR_BG, end_color=COLOR_FK_HDR_BG, fill_type='solid')
FILL_ALT_ROW = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid')

# 对齐
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_TITLE = Alignment(horizontal='left', vertical='center')

# 边框
THIN_SIDE = Side(style='thin', color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BOTTOM_SIDE = Side(style='medium', color=COLOR_TITLE_BG)
BORDER_BOTTOM = Border(bottom=BOTTOM_SIDE)

# ============================================================
# 解析逻辑
# ============================================================

def read_md(filepath):
    with open(filepath, 'rb') as f:
        data = f.read()
    text = data.decode('utf-8', errors='replace')
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    return text


def find_chapter_boundaries(lines):
    chapters = {}
    ch1 = ch2 = ch3 = None
    for i, line in enumerate(lines):
        s = line.strip()
        if s.startswith('# 第一章'):
            ch1 = i
        elif s.startswith('# 第二章'):
            ch2 = i
        elif s.startswith('# 第三章'):
            ch3 = i
    if ch1 is not None:
        chapters['第一章'] = (ch1, ch2 if ch2 else len(lines))
    if ch2 is not None:
        chapters['第二章'] = (ch2, ch3 if ch3 else len(lines))
    if ch3 is not None:
        chapters['第三章'] = (ch3, len(lines))
    return chapters


def find_subdomain_boundaries(lines, start, end):
    subdomains = []
    for i in range(start, end):
        s = lines[i].strip()
        m = re.match(r'^#{1,2}\s+([一二三四五六七八九十]+、.+)', s)
        if m:
            subdomains.append((m.group(1).strip(), i))
    result = []
    for idx, (name, line_no) in enumerate(subdomains):
        next_line = subdomains[idx + 1][1] if idx + 1 < len(subdomains) else end
        result.append((name, line_no, next_line))
    return result


def parse_tables(lines, start, end):
    """解析范围内的所有表定义，返回表列表"""
    tables = []
    i = start

    while i < end:
        line = lines[i].strip()

        # 匹配表标题: ### 1.1 pm_project -- 项目主表 或 ### act_evt_log
        m_std = re.match(r'^#{3,5}\s+[\d.]*\s*(\S+)\s*(?:--\s*(.+))?', line)
        if not m_std or not is_valid_table_name(m_std.group(1)):
            i += 1
            continue

        table_name = m_std.group(1)
        table_comment = m_std.group(2) or ''
        i += 1

        # 解析属性表
        obj_type, biz_meaning, data_count, data_size = '', '', '', ''
        while i < end:
            l = lines[i].strip()
            if l.startswith('|') and '属性' in l and '值' in l:
                i += 1  # skip header
                i += 1  # skip separator
                while i < end and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    parts = [p.strip() for p in lines[i].strip().split('|')]
                    parts = [p for p in parts if p]
                    if len(parts) >= 2:
                        key, val = parts[0], parts[1]
                        if '对象类型' in key:
                            obj_type = val
                        elif '业务含义' in key:
                            biz_meaning = val
                        elif '数据量' in key:
                            data_count = val
                        elif '数据大小' in key or '存储大小' in key:
                            data_size = val
                    i += 1
                break
            elif l.startswith('**字段列表**') or l.startswith('| 字段名'):
                break
            else:
                i += 1

        # 过滤
        if should_filter(table_name, obj_type):
            # 跳到下一个表
            while i < end:
                l = lines[i].strip()
                if re.match(r'^#{2,5}\s+[\d.]*\s*\S+', l) or l == '---':
                    break
                i += 1
            continue

        # 解析字段列表
        fields = []
        ref_table = ''  # "与 xxx 结构相同" 引用
        while i < end:
            l = lines[i].strip()
            if l.startswith('| 字段名'):
                i += 1  # skip header
                i += 1  # skip separator
                while i < end and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    raw_line = lines[i].strip()
                    # 按 | 分割，保留空列以确定列位置
                    # 格式: | col1 | col2 | col3 | col4 | col5 | col6 | col7 |
                    all_parts = raw_line.split('|')
                    # all_parts: ['', ' col1 ', ' col2 ', ..., '']
                    # 去掉首尾空元素，保留中间的
                    if all_parts and all_parts[0].strip() == '':
                        all_parts = all_parts[1:]
                    if all_parts and all_parts[-1].strip() == '':
                        all_parts = all_parts[:-1]

                    # 标准MD表格有7列数据
                    # 列: 字段名(0), 数据类型(1), 可空(2), 默认值(3), 约束(4), 字段描述(5), 业务含义(6)
                    # 但字段描述/业务含义中可能包含 | 导致列数增多
                    # 策略：前5列固定取，第6列到倒数第2列合并为字段描述，最后1列为业务含义

                    stripped = [p.strip() for p in all_parts]
                    col_count = len(stripped)

                    if col_count >= 7:
                        # 标准或含|的情况
                        fname = stripped[0]
                        dtype = stripped[1]
                        nullable = stripped[2]
                        default_val = stripped[3]
                        constraint = stripped[4]
                        # 第6列到倒数第2列合并为字段描述，最后1列为业务含义
                        comment = ' | '.join(stripped[5:-1])
                        biz = stripped[-1]
                        fields.append({
                            '字段名': fname, '数据类型': dtype, '可空': nullable,
                            '默认值': default_val, '约束': constraint, '字段描述': comment,
                            '业务含义': biz,
                        })
                    elif col_count == 6:
                        # 可能是约束列为空被合并，或业务含义为空
                        # 检查第5列是否像约束（PRI/UNI/MUL/auto_increment）
                        fifth = stripped[4]
                        if fifth in ('PRI', 'UNI', 'MUL', 'AUTO_INCREMENT') or 'auto_increment' in fifth.lower() or fifth == '-' or fifth == '':
                            # 第5列是约束，第6列是字段描述，业务含义为空
                            fname = stripped[0]
                            dtype = stripped[1]
                            nullable = stripped[2]
                            default_val = stripped[3]
                            constraint = fifth
                            comment = stripped[5]
                            biz = ''
                        else:
                            # 第5列是字段描述，业务含义为空，约束为空
                            fname = stripped[0]
                            dtype = stripped[1]
                            nullable = stripped[2]
                            default_val = stripped[3]
                            constraint = ''
                            comment = ' | '.join(stripped[4:-1])
                            biz = stripped[-1] if stripped[-1] != stripped[4] else ''
                        fields.append({
                            '字段名': fname, '数据类型': dtype, '可空': nullable,
                            '默认值': default_val, '约束': constraint, '字段描述': comment,
                            '业务含义': biz,
                        })
                    elif col_count == 5:
                        # 约束和业务含义为空
                        fname = stripped[0]
                        dtype = stripped[1]
                        nullable = stripped[2]
                        default_val = stripped[3]
                        constraint = ''
                        # 第5列可能是字段描述
                        comment = stripped[4]
                        biz = ''
                        fields.append({
                            '字段名': fname, '数据类型': dtype, '可空': nullable,
                            '默认值': default_val, '约束': constraint, '字段描述': comment,
                            '业务含义': biz,
                        })
                    elif col_count >= 1:
                        fname = stripped[0]
                        fields.append({
                            '字段名': fname, '数据类型': '', '可空': '',
                            '默认值': '', '约束': '', '字段描述': '',
                            '业务含义': '',
                        })
                    i += 1
                break
            elif l.startswith('**索引') or l.startswith('**外键') or l.startswith('---') or re.match(r'^#{2,5}', l):
                break
            else:
                # 检测 "与 xxx 结构相同/类似" 引用
                m_ref = re.match(r'^与\s+(\S+)\s+结构(相同|类似|基本相同)', l)
                if m_ref:
                    ref_table = m_ref.group(1)
                i += 1

        # 解析索引列表
        indexes = []
        while i < end:
            l = lines[i].strip()
            if l.startswith('| 索引名') and '索引类型' in l:
                i += 1  # skip header
                i += 1  # skip separator
                while i < end and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    non_empty = [p.strip() for p in lines[i].strip().split('|') if p.strip()]
                    if len(non_empty) >= 4:
                        # 前3列固定，后面合并为索引字段
                        idx_name = non_empty[0]
                        idx_type = non_empty[1]
                        uniqueness = non_empty[2]
                        idx_columns = ' | '.join(non_empty[3:])
                        indexes.append({
                            '索引名': idx_name, '索引类型': idx_type,
                            '唯一性': uniqueness, '索引字段': idx_columns,
                        })
                    i += 1
                break
            elif l.startswith('**外键') or l.startswith('---') or re.match(r'^#{2,5}', l):
                break
            else:
                i += 1

        # 解析外键列表
        foreign_keys = []
        while i < end:
            l = lines[i].strip()
            if l.startswith('| 外键名') and '本表字段' in l:
                i += 1  # skip header
                i += 1  # skip separator
                while i < end and lines[i].strip().startswith('|') and '---' not in lines[i]:
                    parts = [p.strip() for p in lines[i].strip().split('|')]
                    parts = [p for p in parts if p]
                    if len(parts) >= 4:
                        foreign_keys.append({
                            '外键名': parts[0], '本表字段': parts[1],
                            '引用表': parts[2], '引用字段': parts[3],
                        })
                    i += 1
                break
            elif l.startswith('---') or re.match(r'^#{2,5}', l):
                break
            else:
                i += 1

        tables.append({
            'table_name': table_name,
            'table_comment': table_comment,
            'obj_type': obj_type or 'BASE TABLE',
            'biz_meaning': biz_meaning,
            'data_count': data_count,
            'data_size': data_size,
            'fields': fields,
            'indexes': indexes,
            'foreign_keys': foreign_keys,
            'ref_table': ref_table,
        })

        # 跳到下一个分隔线或标题
        while i < end:
            l = lines[i].strip()
            if l == '---' or re.match(r'^#{2,5}\s+', l):
                break
            i += 1

    return tables


def is_valid_table_name(name):
    if not name or len(name) < 2:
        return False
    if re.search(r'[\u4e00-\u9fff]', name):
        return False
    if name.startswith('**'):
        return False
    return True


# ============================================================
# Excel 写入逻辑
# ============================================================

# 列宽定义
COL_WIDTHS = {
    'A': 28,   # 字段名
    'B': 18,   # 数据类型
    'C': 8,    # 可空
    'D': 16,   # 默认值
    'E': 16,   # 约束
    'F': 32,   # 字段描述
    'G': 45,   # 业务含义
}


def write_table_to_sheet(ws, table, start_row):
    """将一个表写入工作表，返回下一个可用行号"""
    row = start_row

    # ---- 1. 表名标题行 ----
    title_text = table['table_name']
    if table['table_comment']:
        title_text += f"  ——  {table['table_comment']}"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value=title_text)
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    cell.alignment = ALIGN_TITLE
    cell.border = BORDER_ALL
    for c in range(2, 8):
        ws.cell(row=row, column=c).fill = FILL_TITLE
        ws.cell(row=row, column=c).border = BORDER_ALL
    row += 1

    # ---- 2. 属性区 ----
    attrs = [
        ('对象类型', table['obj_type']),
        ('业务含义', table['biz_meaning']),
        ('数据量', table['data_count']),
        ('数据大小', table['data_size']),
    ]
    for key, val in attrs:
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
        cell_key = ws.cell(row=row, column=1, value='')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell_key = ws.cell(row=row, column=1, value=key)
        cell_key.font = FONT_ATTR_KEY
        cell_key.fill = FILL_ATTR_KEY
        cell_key.alignment = ALIGN_CENTER
        cell_key.border = BORDER_ALL
        ws.cell(row=row, column=2).fill = FILL_ATTR_KEY
        ws.cell(row=row, column=2).border = BORDER_ALL

        cell_val = ws.cell(row=row, column=3, value=val)
        cell_val.font = FONT_ATTR_VAL
        cell_val.fill = FILL_ATTR_VAL
        cell_val.alignment = ALIGN_LEFT
        cell_val.border = BORDER_ALL
        for c in range(4, 8):
            ws.cell(row=row, column=c).fill = FILL_ATTR_VAL
            ws.cell(row=row, column=c).border = BORDER_ALL
        row += 1

    # ---- 3. 字段列表 ----
    field_headers = ['字段名', '数据类型', '可空', '默认值', '约束', '字段描述', '业务含义']
    for col_idx, hdr in enumerate(field_headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=hdr)
        cell.font = FONT_FIELD_HDR
        cell.fill = FILL_FIELD_HDR
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
    row += 1

    for f_idx, field in enumerate(table['fields']):
        row_data = [
            field.get('字段名', ''),
            field.get('数据类型', ''),
            field.get('可空', ''),
            field.get('默认值', ''),
            field.get('约束', ''),
            field.get('字段描述', ''),
            field.get('业务含义', ''),
        ]
        is_alt = f_idx % 2 == 1
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = FONT_FIELD_DATA
            cell.alignment = ALIGN_LEFT if col_idx in (1, 6, 7) else ALIGN_CENTER
            cell.border = BORDER_ALL
            if is_alt:
                cell.fill = FILL_ALT_ROW
        row += 1

    # ---- 4. 索引列表 ----
    if table['indexes']:
        row += 1  # 空行
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value='索引列表')
        cell.font = FONT_SECTION
        cell.alignment = ALIGN_LEFT
        row += 1

        index_headers = ['索引名', '索引类型', '唯一性', '索引字段', '', '', '']
        for col_idx, hdr in enumerate(index_headers[:4], 1):
            cell = ws.cell(row=row, column=col_idx, value=hdr)
            cell.font = FONT_INDEX_HDR
            cell.fill = FILL_INDEX_HDR
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        # 合并后3列
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        row += 1

        for i_idx, idx in enumerate(table['indexes']):
            is_alt = i_idx % 2 == 1
            row_data = [
                idx.get('索引名', ''),
                idx.get('索引类型', ''),
                idx.get('唯一性', ''),
                idx.get('索引字段', ''),
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = FONT_INDEX_DATA
                cell.alignment = ALIGN_LEFT if col_idx == 4 else ALIGN_CENTER
                cell.border = BORDER_ALL
                if is_alt:
                    cell.fill = FILL_ALT_ROW
            # 合并索引字段列
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
            row += 1

    # ---- 5. 外键列表 ----
    if table['foreign_keys']:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value='外键列表')
        cell.font = FONT_SECTION
        cell.alignment = ALIGN_LEFT
        row += 1

        fk_headers = ['外键名', '本表字段', '引用表', '引用字段']
        for col_idx, hdr in enumerate(fk_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=hdr)
            cell.font = FONT_FK_HDR
            cell.fill = FILL_FK_HDR
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        row += 1

        for fk_idx, fk in enumerate(table['foreign_keys']):
            is_alt = fk_idx % 2 == 1
            row_data = [
                fk.get('外键名', ''),
                fk.get('本表字段', ''),
                fk.get('引用表', ''),
                fk.get('引用字段', ''),
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = FONT_FK_DATA
                cell.alignment = ALIGN_LEFT
                cell.border = BORDER_ALL
                if is_alt:
                    cell.fill = FILL_ALT_ROW
            row += 1

    # 表间间隔
    row += 2
    return row


def main():
    print("=" * 60)
    print("导出 database_dict.xlsx（专业格式）")
    print("过滤: temp_*/tmp_* 表 + 视图")
    print("=" * 60)

    # 1. 读取MD
    print("\n1. 读取 MD 文件...")
    text = read_md(MD_FILE)
    lines = text.split('\n')
    print(f"   总行数: {len(lines)}")

    # 2. 识别章节
    print("\n2. 识别章节边界...")
    chapters = find_chapter_boundaries(lines)

    # 3. 解析各业务域
    print("\n3. 解析各业务域表结构...")
    tables_by_domain = {}

    # 先收集所有表到全局字典，用于引用解析
    all_tables_global = {}

    for domain_name, chapter_key, subdomain_keywords in DOMAINS:
        if chapter_key not in chapters:
            continue
        ch_start, ch_end = chapters[chapter_key]
        subdomains = find_subdomain_boundaries(lines, ch_start, ch_end)

        target_subdomains = []
        for sd_name, sd_start, sd_end in subdomains:
            for kw in subdomain_keywords:
                if kw in sd_name or sd_name.startswith(kw):
                    target_subdomains.append((sd_name, sd_start, sd_end))
                    break

        if not target_subdomains:
            target_subdomains = [(chapter_key, ch_start, ch_end)]

        all_tables = []
        for sd_name, sd_start, sd_end in target_subdomains:
            tables = parse_tables(lines, sd_start, sd_end)
            all_tables.extend(tables)

        # 保留MD文件中的章节顺序，不排序
        tables_by_domain[domain_name] = all_tables

        # 加入全局字典
        for t in all_tables:
            all_tables_global[t['table_name']] = t

        print(f"   {domain_name}: {len(all_tables)} 张表")

    # 4. 解析 "与 xxx 结构相同" 引用
    print("\n4. 解析结构引用...")
    ref_resolved = 0
    for domain_name, tables in tables_by_domain.items():
        for t in tables:
            if not t['fields'] and t.get('ref_table'):
                ref_name = t['ref_table']
                if ref_name in all_tables_global and all_tables_global[ref_name]['fields']:
                    t['fields'] = copy.deepcopy(all_tables_global[ref_name]['fields'])
                    t['indexes'] = copy.deepcopy(all_tables_global[ref_name]['indexes'])
                    ref_resolved += 1
                    print(f"   [引用解析] {t['table_name']} <- {ref_name}")
                else:
                    print(f"   [引用未找到] {t['table_name']} -> {ref_name}")

    # 5. 查库补全仍无字段的表
    empty_tables = []
    for domain_name, tables in tables_by_domain.items():
        for t in tables:
            if not t['fields']:
                empty_tables.append((domain_name, t))

    if empty_tables:
        print(f"\n5. 查库补全 {len(empty_tables)} 张空字段表...")
        try:
            conn = pymysql.connect(
                host='localhost', user='root', password='!Q@W3e4r',
                database='dppms_d365', charset='utf8mb4'
            )
            cursor = conn.cursor()

            for domain_name, t in empty_tables:
                table_name = t['table_name']
                # 查字段
                cursor.execute(
                    "SELECT COLUMN_NAME, COLUMN_TYPE, IS_NULLABLE, COLUMN_DEFAULT, "
                    "COLUMN_KEY, EXTRA, COLUMN_COMMENT "
                    "FROM INFORMATION_SCHEMA.COLUMNS "
                    "WHERE TABLE_SCHEMA='dppms_d365' AND TABLE_NAME=%s "
                    "ORDER BY ORDINAL_POSITION", (table_name,)
                )
                rows = cursor.fetchall()
                if rows:
                    for row in rows:
                        col_name, col_type, is_nullable, col_default, col_key, extra, col_comment = row
                        constraint = ''
                        if col_key == 'PRI':
                            constraint = 'PRI'
                        elif col_key == 'UNI':
                            constraint = 'UNI'
                        elif col_key == 'MUL':
                            constraint = 'MUL'
                        if extra and 'auto_increment' in extra:
                            constraint += ', AUTO_INCREMENT' if constraint else 'AUTO_INCREMENT'
                        default_val = str(col_default) if col_default is not None else '-'
                        if default_val == 'None':
                            default_val = '-'
                        t['fields'].append({
                            '字段名': col_name,
                            '数据类型': col_type,
                            '可空': is_nullable,
                            '默认值': default_val,
                            '约束': constraint,
                            '字段描述': col_comment or '',
                            '业务含义': '',
                        })
                    print(f"   [查库补全] {table_name}: {len(rows)} 个字段")
                else:
                    print(f"   [查库无数据] {table_name}")

                # 查索引
                cursor.execute(
                    "SELECT INDEX_NAME, INDEX_TYPE, NON_UNIQUE, "
                    "GROUP_CONCAT(COLUMN_NAME ORDER BY SEQ_IN_INDEX) AS COLUMNS "
                    "FROM INFORMATION_SCHEMA.STATISTICS "
                    "WHERE TABLE_SCHEMA='dppms_d365' AND TABLE_NAME=%s "
                    "GROUP BY INDEX_NAME, INDEX_TYPE, NON_UNIQUE",
                    (table_name,)
                )
                idx_rows = cursor.fetchall()
                if idx_rows and not t['indexes']:
                    for row in idx_rows:
                        idx_name, idx_type, non_unique, idx_columns = row
                        t['indexes'].append({
                            '索引名': idx_name,
                            '索引类型': idx_type,
                            '唯一性': 'UNIQUE' if not non_unique else 'NON-UNIQUE',
                            '索引字段': idx_columns,
                        })

            cursor.close()
            conn.close()
        except Exception as e:
            print(f"   [数据库连接失败] {e}")
    else:
        print("\n5. 所有表均有字段数据，无需查库补全")

    # 6. 写入 Excel
    print("\n6. 写入 Excel...")
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for domain_name, tables in tables_by_domain.items():
        if not tables:
            continue

        sheet_name = domain_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 设置列宽
        for col_letter, width in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        row = 1
        for table in tables:
            row = write_table_to_sheet(ws, table, row)

        # 冻结首行
        ws.freeze_panes = 'A2'

        print(f"   {domain_name}: {len(tables)} 张表, {row} 行")

    wb.save(XLSX_FILE)
    print(f"\n7. Excel 已保存: {XLSX_FILE}")

    # 统计
    total_tables = sum(len(t) for t in tables_by_domain.values())
    total_fields = sum(len(t['fields']) for tables in tables_by_domain.values() for t in tables)
    total_indexes = sum(len(t['indexes']) for tables in tables_by_domain.values() for t in tables)
    total_fks = sum(len(t['foreign_keys']) for tables in tables_by_domain.values() for t in tables)

    # 字段描述和业务含义覆盖率
    has_comment = sum(1 for tables in tables_by_domain.values() for t in tables for f in t['fields'] if f.get('字段描述') and f['字段描述'] not in ('-', 'None', ''))
    has_biz = sum(1 for tables in tables_by_domain.values() for t in tables for f in t['fields'] if f.get('业务含义') and f['业务含义'] not in ('-', 'None', '', '业务含义待确认'))
    empty_both = sum(1 for tables in tables_by_domain.values() for t in tables for f in t['fields'] if (not f.get('字段描述') or f['字段描述'] in ('-', 'None', '')) and (not f.get('业务含义') or f['业务含义'] in ('-', 'None', '', '业务含义待确认')))

    print(f"\n统计:")
    print(f"  总表数(已过滤temp/tmp/视图): {total_tables}")
    print(f"  总字段数: {total_fields}")
    print(f"  总索引数: {total_indexes}")
    print(f"  总外键数: {total_fks}")
    print(f"  有字段描述: {has_comment} ({has_comment*100//total_fields}%)")
    print(f"  有业务含义: {has_biz} ({has_biz*100//total_fields}%)")
    print(f"  两者都为空: {empty_both} ({empty_both*100//total_fields}%)")

    # 列出两者都为空的表（前10个）
    empty_tables = []
    for domain_name, tables in tables_by_domain.items():
        for t in tables:
            empty_count = sum(1 for f in t['fields'] if (not f.get('字段描述') or f['字段描述'] in ('-', 'None', '')) and (not f.get('业务含义') or f['业务含义'] in ('-', 'None', '', '业务含义待确认')))
            if empty_count > 0:
                empty_tables.append((t['table_name'], empty_count, len(t['fields'])))
    empty_tables.sort(key=lambda x: -x[1])
    if empty_tables:
        print(f"\n  字段描述和业务含义都为空最多的表（前10个）:")
        for tname, ec, tc in empty_tables[:10]:
            print(f"    {tname}: {ec}/{tc} 字段为空")

    # 列出各域表名
    for domain_name, tables in tables_by_domain.items():
        names = [t['table_name'] for t in tables]
        print(f"\n  [{domain_name}] ({len(tables)} 张): {', '.join(names[:10])}{'...' if len(names) > 10 else ''}")


if __name__ == '__main__':
    main()
