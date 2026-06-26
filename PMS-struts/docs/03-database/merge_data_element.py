#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
合并脚本：
以 database_dict_flat_final_match.xlsx 为基表，
从 项目管理表数据元.xlsx 中按表名+字段名匹配追加9列新数据。
不存在的字段按顺序插入基表，保持基表样式。
"""
import re, os, copy
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_XLSX = os.path.join(BASE_DIR, 'database_dict_flat_final_match.xlsx')
MATCH_XLSX = os.path.join(BASE_DIR, '项目管理表数据元.xlsx')
OUT_XLSX = os.path.join(BASE_DIR, 'database_dict_flat_final_merged_v2.xlsx')

# 新增的9列
NEW_COLS = ['来源', '对应业务', '数据元类型', '数据现状', '处理方案', '数据资产', '关联资产', '应用主题', '自动回传']
NEW_COL_START = 10  # 从第J列开始

# ============================================================
# 样式定义（与 export_flat.py 一致）
# ============================================================
COLOR_TITLE_BG = '1F4E79'
COLOR_TITLE_FONT = 'FFFFFF'
COLOR_HDR_BG = '2E75B6'
COLOR_HDR_FONT = 'FFFFFF'
COLOR_ALT_ROW = 'F2F7FB'
COLOR_BORDER = 'B4C6E7'

FONT_HDR = Font(name='微软雅黑', bold=True, size=10, color=COLOR_HDR_FONT)
FONT_DATA = Font(name='微软雅黑', size=9, color='333333')

FILL_HDR = PatternFill(start_color=COLOR_HDR_BG, end_color=COLOR_HDR_BG, fill_type='solid')
FILL_ALT = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type='solid')

ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN_SIDE = Side(style='thin', color=COLOR_BORDER)
BORDER_ALL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

# 列宽
COL_WIDTHS = {
    'A': 30, 'B': 28, 'C': 22, 'D': 18, 'E': 8, 'F': 14, 'G': 14,
    'H': 32, 'I': 40,
    'J': 12, 'K': 18, 'L': 12, 'M': 14, 'N': 18, 'O': 12, 'P': 12, 'Q': 14, 'R': 12,
}

HEADERS = ['表名', '表描述', '字段名', '数据类型', '可空', '默认值', '约束', '字段描述', '业务含义'] + NEW_COLS


# ============================================================
# 1. 解析匹配表
# ============================================================
def parse_match_xlsx(filepath):
    """
    解析项目管理表数据元.xlsx
    返回: {table_name: [match_field, ...]}
    match_field = {
        'fname': str or '',  # 字段名（可能为空）
        'dtype': str, 'nullable': str, 'default_val': str, 'constraint': str,
        'comment': str, 'biz': str,
        '来源': str, '对应业务': str, '数据元类型': str, '数据现状': str,
        '处理方案': str, '数据资产': str, '关联资产': str, '应用主题': str, '自动回传': str,
    }
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    tables = {}
    current_table = None
    current_tdesc = ''
    in_field_section = False

    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c6 = ws.cell(r, 6).value

        # 表名行：C1包含"——"或只有表名（无C2）
        if c1 and c2 is None and str(c1) != '字段名' and str(c1) not in ('对象类型', '业务含义', '数据量', '数据大小', '索引列表', '索引名'):
            s1 = str(c1).strip()
            if '——' in s1:
                parts = s1.split('——')
                current_table = parts[0].strip()
                current_tdesc = parts[1].strip() if len(parts) > 1 else ''
            elif re.match(r'^[a-zA-Z_]\w*$', s1):
                current_table = s1
                current_tdesc = ''
            else:
                continue
            tables[current_table] = []
            in_field_section = False
            continue

        # 属性行（对象类型/业务含义等），跳过
        if c1 and str(c1).strip() in ('对象类型', '业务含义', '数据量', '数据大小'):
            continue

        # 索引列表，跳过后续直到下一个表
        if c1 and str(c1).strip() == '索引列表':
            in_field_section = False
            continue

        # 字段表头行
        if c1 and str(c1).strip() == '字段名':
            in_field_section = True
            continue

        # 字段数据行
        if in_field_section and current_table:
            fname = str(c1).strip() if c1 else ''
            dtype = str(c2).strip() if c2 else ''
            nullable = str(c3).strip() if c3 else ''
            default_val = str(ws.cell(r, 4).value or '').strip()
            constraint = str(ws.cell(r, 5).value or '').strip()
            comment = str(c6).strip() if c6 else ''
            biz = str(ws.cell(r, 7).value or '').strip()

            # 9个新列
            new_data = {}
            for idx, col_name in enumerate(NEW_COLS):
                new_data[col_name] = str(ws.cell(r, 8 + idx).value or '').strip()

            # 跳过完全空的行
            if not fname and not comment and not biz:
                # 检查是否有新列数据
                has_new = any(new_data[c] for c in NEW_COLS)
                if not has_new:
                    continue

            field = {
                'fname': fname,
                'dtype': dtype,
                'nullable': nullable,
                'default_val': default_val,
                'constraint': constraint,
                'comment': comment,
                'biz': biz,
            }
            field.update(new_data)
            tables[current_table].append(field)

    wb.close()
    return tables


# ============================================================
# 2. 解析基表
# ============================================================
def parse_base_xlsx(filepath):
    """
    解析 database_dict_flat_final_match.xlsx
    返回: {sheet_name: [base_table, ...]}
    base_table = {
        'table_name': str, 'table_desc': str,
        'fields': [base_field, ...]
    }
    base_field = {
        'fname': str, 'dtype': str, 'nullable': str, 'default_val': str,
        'constraint': str, 'comment': str, 'biz': str,
    }
    """
    wb = openpyxl.load_workbook(filepath)
    result = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        tables = []
        current_table = None
        current_fields = []

        for r in range(1, ws.max_row + 1):
            a = ws.cell(r, 1).value  # 表名
            b = ws.cell(r, 2).value  # 表描述
            c = ws.cell(r, 3).value  # 字段名

            # 表头行
            if a and str(a).strip() == '表名':
                # 保存上一个表
                if current_table and current_fields:
                    current_table['fields'] = current_fields
                    tables.append(current_table)
                current_table = None
                current_fields = []
                continue

            # 空行 = 表间分隔
            if a is None and c is None:
                if current_table and current_fields:
                    current_table['fields'] = current_fields
                    tables.append(current_table)
                current_table = None
                current_fields = []
                continue

            # 数据行
            if a and c:
                if current_table is None:
                    current_table = {
                        'table_name': str(a).strip(),
                        'table_desc': str(b or '').strip(),
                    }
                current_fields.append({
                    'fname': str(c).strip(),
                    'dtype': str(ws.cell(r, 4).value or '').strip(),
                    'nullable': str(ws.cell(r, 5).value or '').strip(),
                    'default_val': str(ws.cell(r, 6).value or '').strip(),
                    'constraint': str(ws.cell(r, 7).value or '').strip(),
                    'comment': str(ws.cell(r, 8).value or '').strip(),
                    'biz': str(ws.cell(r, 9).value or '').strip(),
                })

        # 最后一个表
        if current_table and current_fields:
            current_table['fields'] = current_fields
            tables.append(current_table)

        result[sn] = tables

    wb.close()
    return result


# ============================================================
# 3. 合并逻辑
# ============================================================
def merge_tables(base_tables, match_data):
    """
    对基表中的每个表，匹配追加9列新数据。
    匹配逻辑：
    - 按表名+字段名精确匹配
    - 匹配表中不存在字段名的行为新增字段
    - 新增字段按匹配表所在字段位置顺序插入到对应字段之后
    - 原有列内容不做修改，只追加新列数据
    """
    merged_tables = []

    for bt in base_tables:
        tname = bt['table_name']
        tdesc = bt['table_desc']
        base_fields = bt['fields']

        # 构建基表字段名集合
        base_fname_set = set(f['fname'] for f in base_fields if f['fname'])

        # 获取匹配表中该表的数据
        match_fields = match_data.get(tname, [])

        if not match_fields:
            # 没有匹配数据，原样保留，新列留空
            new_fields = []
            for f in base_fields:
                nf = dict(f)
                for c in NEW_COLS:
                    nf[c] = ''
                new_fields.append(nf)
            merged_tables.append({
                'table_name': tname,
                'table_desc': tdesc,
                'fields': new_fields,
            })
            continue

        # 构建匹配字段的查找索引：fname -> mf
        match_by_fname = {}
        for mf in match_fields:
            if mf['fname']:
                match_by_fname[mf['fname']] = mf

        # ============================================================
        # 核心逻辑：遍历匹配表字段，按位置关系构建最终字段列表
        # - 有字段名且在基表中存在 → 匹配，追加新列数据
        # - 有字段名但不在基表中 → 新增字段，插入到当前位置
        # - 无字段名（只有字段描述）→ 新增字段，插入到前一个字段之后
        # ============================================================

        # 先给基表所有字段加上空的新列
        base_with_new = []
        for f in base_fields:
            nf = dict(f)
            for c in NEW_COLS:
                nf[c] = ''
            base_with_new.append(nf)

        # 构建基表字段名到索引的映射
        base_fname_to_idx = {}
        for idx, f in enumerate(base_fields):
            if f['fname']:
                base_fname_to_idx[f['fname']] = idx

        # 遍历匹配表字段，按位置关系处理
        # 记录每个基表字段之后需要插入的新字段
        # key = 基表字段名, value = [要插入在该字段之后的匹配字段列表]
        insert_after = {}  # fname -> [match_field, ...]
        # 记录匹配表开头（在第一个匹配字段之前）需要插入的字段
        insert_before_first = []

        last_matched_fname = None  # 上一个匹配到的基表字段名

        for mf in match_fields:
            mf_fname = mf['fname']

            if mf_fname and mf_fname in base_fname_set:
                # 有字段名且在基表中存在 → 精确匹配
                # 追加新列数据到基表对应字段
                base_idx = base_fname_to_idx[mf_fname]
                for c in NEW_COLS:
                    if mf.get(c):
                        base_with_new[base_idx][c] = mf[c]
                last_matched_fname = mf_fname

            elif mf_fname or mf.get('comment'):
                # 有字段名但不在基表中，或无字段名（只有字段描述）→ 新增字段
                new_f = {
                    'fname': mf.get('fname', ''),
                    'dtype': mf.get('dtype', ''),
                    'nullable': mf.get('nullable', ''),
                    'default_val': mf.get('default_val', ''),
                    'constraint': mf.get('constraint', ''),
                    'comment': mf.get('comment', ''),
                    'biz': mf.get('biz', ''),
                }
                for c in NEW_COLS:
                    new_f[c] = mf.get(c, '')

                if last_matched_fname:
                    # 插入到上一个匹配字段之后
                    if last_matched_fname not in insert_after:
                        insert_after[last_matched_fname] = []
                    insert_after[last_matched_fname].append(new_f)
                else:
                    # 还没有匹配到任何基表字段，插入到开头
                    insert_before_first.append(new_f)

        # 构建最终字段列表：按基表顺序，在每个字段后插入新增字段
        final_fields = []

        # 开头的新增字段
        final_fields.extend(insert_before_first)

        for f in base_with_new:
            final_fields.append(f)
            fname = f.get('fname', '')
            if fname in insert_after:
                final_fields.extend(insert_after[fname])

        merged_tables.append({
            'table_name': tname,
            'table_desc': tdesc,
            'fields': final_fields,
        })

    return merged_tables


# ============================================================
# 4. 写入 Excel
# ============================================================
def write_merged_xlsx(filepath, all_sheets_data):
    """写入合并后的Excel，保持基表样式"""
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, tables in all_sheets_data.items():
        if not tables:
            continue
        sn = sheet_name[:31]
        ws = wb.create_sheet(title=sn)

        # 列宽
        for col_letter, width in COL_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        row = 1
        num_cols = len(HEADERS)

        for table in tables:
            tname = table['table_name']
            tdesc = table['table_desc']

            # 列名表头行
            for col_idx, hdr in enumerate(HEADERS, 1):
                cell = ws.cell(row=row, column=col_idx, value=hdr)
                cell.font = FONT_HDR
                cell.fill = FILL_HDR
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_ALL
            row += 1

            # 字段数据行
            for f_idx, field in enumerate(table['fields']):
                is_alt = f_idx % 2 == 1
                row_data = [
                    tname,
                    tdesc,
                    field.get('fname', ''),
                    field.get('dtype', ''),
                    field.get('nullable', ''),
                    field.get('default_val', ''),
                    field.get('constraint', ''),
                    field.get('comment', ''),
                    field.get('biz', ''),
                ]
                # 9列新数据
                for c in NEW_COLS:
                    row_data.append(field.get(c, ''))

                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col_idx, value=val if val else '')
                    cell.font = FONT_DATA
                    if col_idx in (1, 2, 3, 8, 9):
                        cell.alignment = ALIGN_LEFT
                    else:
                        cell.alignment = ALIGN_CENTER
                    cell.border = BORDER_ALL
                    if is_alt:
                        cell.fill = FILL_ALT
                row += 1

            # 空行（表间分隔）
            row += 1

        ws.freeze_panes = 'A2'

    wb.save(filepath)


# ============================================================
# 主流程
# ============================================================
def main():
    print("=" * 60)
    print("合并数据字典与数据元")
    print("=" * 60)

    # 1. 解析匹配表
    print("\n1. 解析匹配表...")
    match_data = parse_match_xlsx(MATCH_XLSX)
    print(f"   匹配表中共 {len(match_data)} 张表")
    for tname, fields in list(match_data.items())[:5]:
        print(f"   {tname}: {len(fields)} 个字段")

    # 2. 解析基表
    print("\n2. 解析基表...")
    base_data = parse_base_xlsx(BASE_XLSX)
    for sn, tables in base_data.items():
        total_fields = sum(len(t['fields']) for t in tables)
        print(f"   {sn}: {len(tables)} 张表, {total_fields} 字段")

    # 3. 合并
    print("\n3. 合并数据...")
    all_sheets_merged = {}
    match_count = 0
    insert_count = 0

    for sn, tables in base_data.items():
        merged = merge_tables(tables, match_data)

        # 统计
        for t in merged:
            for f in t['fields']:
                has_new = any(f.get(c) for c in NEW_COLS)
                if has_new:
                    match_count += 1
                # 检查是否是插入的字段（字段名不在原始基表中）
                # 简单判断：如果字段名在匹配表中但不在基表中

        all_sheets_merged[sn] = merged
        total_fields = sum(len(t['fields']) for t in merged)
        print(f"   {sn}: {len(merged)} 张表, {total_fields} 字段(含插入)")

    # 统计插入的字段数
    base_pm_tables = {t['table_name']: set(f['fname'] for f in t['fields']) for t in base_data.get('项目管理', [])}
    for t in all_sheets_merged.get('项目管理', []):
        tname = t['table_name']
        base_fnames = base_pm_tables.get(tname, set())
        for f in t['fields']:
            if f['fname'] and f['fname'] not in base_fnames:
                insert_count += 1
            elif not f['fname'] and f.get('comment'):
                # 没有字段名但有字段描述的插入行
                base_comments = set()
                for bt in base_data.get('项目管理', []):
                    if bt['table_name'] == tname:
                        base_comments = set(bf['comment'] for bf in bt['fields'])
                if f['comment'] not in base_comments:
                    insert_count += 1

    print(f"\n   匹配追加新列数据的字段: {match_count}")
    print(f"   从匹配表插入的新字段: {insert_count}")

    # 4. 写入
    print("\n4. 写入 Excel...")
    write_merged_xlsx(OUT_XLSX, all_sheets_merged)
    print(f"   已保存: {OUT_XLSX}")

    # 最终统计
    total_tables = sum(len(t) for t in all_sheets_merged.values())
    total_fields = sum(len(t['fields']) for tables in all_sheets_merged.values() for t in tables)
    print(f"\n统计:")
    print(f"  总表数: {total_tables}")
    print(f"  总字段数: {total_fields}")


if __name__ == '__main__':
    main()
