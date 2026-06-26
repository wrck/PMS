#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证Excel：检查之前空字段的表是否已补全"""
from openpyxl import load_workbook

TARGET_TABLES = [
    'pm_project_property_from_sms',
    'pm_project_property_from_sms_history',
    'pm_project_property_af_from_sms',
    'project_info_from_sms',
    'pm_order_data_from_erp_sap',
    'pm_order_line_from_erp_source',
    'pm_project_product_af_from_sms',
    'pm_project_market_relations_from_sms',
]

wb = load_workbook('d:\\EclipseWorkspace\\Parctice\\PMS\\PMS-struts\\docs\\03-database\\database_dict.xlsx')
for name in wb.sheetnames:
    ws = wb[name]
    row = 1
    while row <= ws.max_row:
        cell = ws.cell(row=row, column=1)
        if cell.font and cell.font.bold and cell.font.size == 12 and cell.font.color and cell.font.color.rgb and 'FFFFFF' in str(cell.font.color.rgb):
            title = str(cell.value or '')
            for target in TARGET_TABLES:
                if target in title:
                    # 找到字段表头
                    field_header_row = row + 5
                    field_count = 0
                    if field_header_row <= ws.max_row and ws.cell(row=field_header_row, column=1).value == '字段名':
                        r = field_header_row + 1
                        while r <= ws.max_row:
                            v = ws.cell(row=r, column=1).value
                            if v and str(v).strip() and not str(v).startswith('索引') and not str(v).startswith('外键'):
                                field_count += 1
                            else:
                                break
                            r += 1
                    print(f'[{name}] {title[:50]}: {field_count} 个字段')
                    break
            row += 1
        else:
            row += 1
