#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""验证 Excel 字段描述和业务含义覆盖率"""
from openpyxl import load_workbook
import os

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX_FILE = os.path.join(BASE, 'database_dict_final_v2.xlsx')

wb = load_workbook(XLSX_FILE, read_only=True)

total_tables = 0
total_fields = 0
empty_comment = 0
empty_biz = 0
has_comment = 0
has_biz = 0

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    in_field_section = False
    table_count = 0
    
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not row[0]:
            continue
        val0 = str(row[0]).strip()
        
        # 检测表名标题行
        if '——' in val0:
            table_count += 1
            in_field_section = False
        
        # 检测字段表头行
        if val0 == '字段名':
            in_field_section = True
            continue
        
        # 字段数据行
        if in_field_section and val0 and val0 not in ('索引名', '外键名'):
            # 检查是否有数据类型（第2列）
            if row[1] and isinstance(row[1], str):
                total_fields += 1
                comment = str(row[5]).strip() if row[5] else ''
                biz = str(row[6]).strip() if row[6] else ''
                if comment and comment not in ('-', 'None', ''):
                    has_comment += 1
                else:
                    empty_comment += 1
                if biz and biz not in ('-', 'None', '', '业务含义待确认'):
                    has_biz += 1
                else:
                    empty_biz += 1
        
        # 离开字段区域
        if val0 in ('索引名', '外键名', '索引列表', '外键列表'):
            in_field_section = False
    
    total_tables += table_count
    print(f"  {ws_name}: {table_count} 张表")

wb.close()

print(f"\n总表数: {total_tables}")
print(f"总字段数: {total_fields}")
print(f"有字段描述: {has_comment} ({has_comment*100//total_fields}%)")
print(f"无字段描述: {empty_comment} ({empty_comment*100//total_fields}%)")
print(f"有业务含义: {has_biz} ({has_biz*100//total_fields}%)")
print(f"无业务含义(含待确认): {empty_biz} ({empty_biz*100//total_fields}%)")
